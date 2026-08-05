# -*- coding: utf-8 -*-
import os, random, time, sys, socket, hashlib, re, json, unicodedata, csv
from io import StringIO
from datetime import datetime, timedelta
from contextlib import contextmanager
from functools import wraps
from urllib.parse import parse_qs, urlparse
from flask import Flask, Response, render_template, request, redirect, url_for, flash, session, jsonify
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge
from werkzeug.utils import secure_filename
import psycopg2
from psycopg2 import IntegrityError
from psycopg2.extras import RealDictCursor

# ===== Sortie console UTF-8 (Windows) =====
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# ===== Flask =====
BASE = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=os.path.join(BASE, "templates"),
             static_folder=os.path.join(BASE, "static"))
app.secret_key = os.getenv("SECRET_KEY", "change-this-secret")
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB

@app.errorhandler(RequestEntityTooLarge)
def handle_upload_too_large(error):
    if request.path.startswith("/admin/results/upload"):
        flash("ملف Excel كبير جداً. يرجى تصديره كملف .xlsx فقط وتقليل حجمه إن أمكن.", "danger")
        return redirect(url_for("admin_dashboard") + "#national-results")
    return error

@app.errorhandler(Exception)
def handle_unexpected_error(error):
    if request.path.startswith("/admin/results/upload"):
        app.logger.exception("Unhandled results upload error")
        flash(f"تعذر معالجة ملف Excel: {error}", "danger")
        return redirect(url_for("admin_dashboard") + "#national-results")
    if isinstance(error, HTTPException):
        return error
    app.logger.exception("Unhandled application error")
    return "Internal Server Error", 500

@app.after_request
def add_api_cors_headers(response):
    if request.path.startswith("/api/"):
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PATCH, DELETE, OPTIONS"
    if request.path in {"/", "/results"} or request.path.startswith("/results/") or request.path.startswith("/api/results/") or request.path.startswith("/api/visitors/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

# Dossiers dâ€™upload
UP = os.path.join(BASE, "static", "uploads")
PAY_DIR = os.path.join(UP, "payments")
VID_DIR = os.path.join(UP, "videos")
PDF_DIR = os.path.join(UP, "pdfs")
OFFERS_DIR = os.path.join(UP, "offers")
for p in (PAY_DIR, VID_DIR, PDF_DIR, OFFERS_DIR):
    os.makedirs(p, exist_ok=True)

IMG_EXT   = {"jpg","jpeg","png","webp","pdf"}  # payment proof allows pdf too
OFFER_IMG_EXT = {"jpg", "jpeg", "png", "webp"}
VIDEO_EXT = {"mp4","webm","mkv","avi","mov","ogg"}
PDF_EXT   = {"pdf"}
EXCEL_EXT = {"xlsx","xlsm"}
def allowed(fn, exts): return "." in fn and fn.rsplit(".",1)[1].lower() in exts

# ===== PostgreSQL =====
DATABASE_URL = os.getenv("DATABASE_URL")
DB = dict(
    host=os.getenv("PGHOST", "localhost"),
    port=os.getenv("PGPORT", "5432"),
    user=os.getenv("PGUSER", "postgres"),
    password=os.getenv("PGPASSWORD", ""),
    dbname=os.getenv("PGDATABASE", "school_app"),
)

def database_dsn():
    if not DATABASE_URL:
        return None
    dsn = DATABASE_URL
    if "render.com" in dsn.lower() and "sslmode=" not in dsn.lower():
        dsn += "&sslmode=require" if "?" in dsn else "?sslmode=require"
    return dsn

@contextmanager
def db():
    conn = psycopg2.connect(database_dsn()) if DATABASE_URL else psycopg2.connect(**DB)
    try:
        yield conn
    finally:
        conn.close()

def dict_cursor(conn):
    return conn.cursor(cursor_factory=RealDictCursor)

def parse_selected_subjects(value):
    if not value:
        return []
    if isinstance(value, list):
        items = value
    else:
        try:
            items = json.loads(value)
        except (TypeError, ValueError):
            items = str(value).split(",")
    return [str(item).strip() for item in items if str(item).strip()]

def selected_subjects_json(subjects, allowed_subjects=None):
    allowed = set(allowed_subjects or [])
    cleaned = []
    for subject in subjects or []:
        item = str(subject).strip()
        if item and item not in cleaned and (not allowed or item in allowed):
            cleaned.append(item)
    return json.dumps(cleaned, ensure_ascii=False)

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()

def remove_upload(folder, filename):
    if not filename:
        return
    path = os.path.abspath(os.path.join(folder, filename))
    folder_abs = os.path.abspath(folder)
    if os.path.commonpath([folder_abs, path]) != folder_abs:
        return
    try:
        os.remove(path)
    except FileNotFoundError:
        pass
    except OSError as e:
        print("Upload cleanup error:", e)

def delete_user_with_teacher_lessons(user_id):
    uploads_to_remove = []
    lessons_deleted = 0
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT id, role FROM users WHERE id=%s", (user_id,))
        user = cur.fetchone()
        if not user:
            cur.close()
            return {"user_deleted": False, "lessons_deleted": 0}

        if user["role"] == "teacher":
            cur.execute("SELECT video_file, pdf_file FROM lessons WHERE uploaded_by=%s", (user_id,))
            lessons = cur.fetchall()
            uploads_to_remove = [
                (lesson.get("video_file"), lesson.get("pdf_file"))
                for lesson in lessons
            ]
            cur.execute("DELETE FROM lessons WHERE uploaded_by=%s", (user_id,))
            lessons_deleted = cur.rowcount

        cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
        user_deleted = cur.rowcount > 0
        conn.commit()
        cur.close()

    for video_file, pdf_file in uploads_to_remove:
        remove_upload(VID_DIR, video_file)
        remove_upload(PDF_DIR, pdf_file)

    return {"user_deleted": user_deleted, "lessons_deleted": lessons_deleted}

def google_drive_preview_url(url):
    if not url:
        return ""
    patterns = [
        r"drive\.google\.com/file/d/([^/]+)",
        r"drive\.google\.com/open\?id=([^&]+)",
        r"drive\.google\.com/uc\?id=([^&]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return f"https://drive.google.com/file/d/{match.group(1)}/preview"
    return url

def video_embed_url(url):
    if not url:
        return ""
    url = url.strip()
    parsed = urlparse(url)
    host = parsed.netloc.lower().removeprefix("www.")

    if host == "youtu.be":
        video_id = parsed.path.strip("/").split("/")[0]
        return f"https://www.youtube.com/embed/{video_id}" if video_id else url

    if host in {"youtube.com", "m.youtube.com", "music.youtube.com", "youtube-nocookie.com"}:
        path_parts = [part for part in parsed.path.split("/") if part]
        if path_parts and path_parts[0] == "embed":
            return url
        if path_parts and path_parts[0] in {"shorts", "live"} and len(path_parts) > 1:
            return f"https://www.youtube.com/embed/{path_parts[1]}"
        video_id = parse_qs(parsed.query).get("v", [None])[0]
        if video_id:
            return f"https://www.youtube.com/embed/{video_id}"

    return google_drive_preview_url(url)

app.jinja_env.filters["video_embed_url"] = video_embed_url

# ===== SMS (Twilio ou mode DEV) =====
TWILIO_SID   = os.getenv("TWILIO_ACCOUNT_SID")
TWILIO_TOKEN = os.getenv("TWILIO_AUTH_TOKEN")
TWILIO_FROM  = os.getenv("TWILIO_FROM_NUMBER")
DEV_SMS      = not (TWILIO_SID and TWILIO_TOKEN and TWILIO_FROM)

def send_sms(phone: str, message: str):
    """
    En DEV on imprime une version ASCII-safe (pas d'accents) pour Ã©viter UnicodeEncodeError,
    et on affiche le texte FR via flash.
    """
    if DEV_SMS:
        import re
        m = re.search(r"(\d{4,6})", message)
        code = m.group(1) if m else "xxxxxx"
        print(f"[DEV SMS] to {phone}: code={code}")  # ASCII only
        try:
            flash(f"(DEV) Code envoyÃ© au {phone} : {code}", "secondary")
        except Exception:
            pass
        return True
    try:
        from twilio.rest import Client
        Client(TWILIO_SID, TWILIO_TOKEN).messages.create(to=phone, from_=TWILIO_FROM, body=message)
        return True
    except Exception as e:
        print("SMS error:", e)
        flash("Ã‰chec dâ€™envoi du SMS. VÃ©rifiez la configuration Twilio.", "danger")
        return False

# ===== OTP =====
def create_otp(phone: str, purpose: str) -> str:
    code = f"{random.randint(100000, 999999)}"
    expires_at = (datetime.utcnow() + timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO verification_codes (phone, code, purpose, expires_at)
                       VALUES (%s,%s,%s,%s)""", (phone, code, purpose, expires_at))
        conn.commit(); cur.close()
    return code

def verify_otp(phone: str, purpose: str, code: str) -> bool:
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT id FROM verification_codes
                       WHERE phone=%s AND purpose=%s AND code=%s
                         AND used=FALSE AND expires_at>CURRENT_TIMESTAMP
                       ORDER BY id DESC LIMIT 1""", (phone, purpose, code))
        row = cur.fetchone()
        if not row:
            cur.close(); return False
        cur2 = conn.cursor()
        cur2.execute("UPDATE verification_codes SET used=TRUE WHERE id=%s", (row["id"],))
        conn.commit(); cur2.close(); cur.close()
    return True

# ===== Auth =====
ADMIN_ROLES = {"admin", "developer"}
FINANCE_ROLES = {"finance", "developer"}
MARKETER_ROLES = {"marketer", "developer"}
FINANCE_USERNAME = os.getenv("FINANCE_USERNAME", "finance")
FINANCE_PHONE = os.getenv("FINANCE_PHONE", "00000000")
FINANCE_PASSWORD = os.getenv("FINANCE_PASSWORD", "finance123")
MARKETER_USERNAME = os.getenv("MARKETER_USERNAME", "المسوق")
MARKETER_PHONE = os.getenv("MARKETER_PHONE", "32324816")
MARKETER_PASSWORD = os.getenv("MARKETER_PASSWORD", "32324816")

def ensure_default_finance_user(cur):
    if not (FINANCE_USERNAME and FINANCE_PHONE and FINANCE_PASSWORD):
        return

    cur.execute("SELECT id FROM users WHERE username=%s", (FINANCE_USERNAME,))
    existing = cur.fetchone()
    if existing:
        cur.execute("""
            UPDATE users
            SET password=%s,
                role='finance',
                status='active',
                phone_verified=TRUE,
                payment_status='not_applicable'
            WHERE username=%s
        """, (hash_password(FINANCE_PASSWORD), FINANCE_USERNAME))
        return

    phone = FINANCE_PHONE
    cur.execute("SELECT id FROM users WHERE phone=%s", (phone,))
    if cur.fetchone():
        phone = f"{FINANCE_USERNAME}-account"
        suffix = 2
        while True:
            cur.execute("SELECT id FROM users WHERE phone=%s", (phone,))
            if not cur.fetchone():
                break
            phone = f"{FINANCE_USERNAME}-account-{suffix}"
            suffix += 1

    cur.execute("""
        INSERT INTO users
            (username, phone, password, role, status, phone_verified, payment_status)
        VALUES (%s,%s,%s,'finance','active',TRUE,'not_applicable')
    """, (FINANCE_USERNAME, phone, hash_password(FINANCE_PASSWORD)))

def ensure_default_marketer_user(cur):
    if not (MARKETER_USERNAME and MARKETER_PHONE and MARKETER_PASSWORD):
        return

    cur.execute("SELECT id FROM users WHERE username=%s", (MARKETER_USERNAME,))
    existing = cur.fetchone()
    if existing:
        cur.execute("""
            UPDATE users
            SET phone=%s,
                password=%s,
                role='marketer',
                status='active',
                phone_verified=TRUE,
                payment_status='not_applicable'
            WHERE username=%s
        """, (MARKETER_PHONE, hash_password(MARKETER_PASSWORD), MARKETER_USERNAME))
        return

    cur.execute("""
        INSERT INTO users
            (username, phone, password, role, status, phone_verified, payment_status)
        VALUES (%s,%s,%s,'marketer','active',TRUE,'not_applicable')
        ON CONFLICT (phone) DO UPDATE
        SET username=EXCLUDED.username,
            password=EXCLUDED.password,
            role='marketer',
            status='active',
            phone_verified=TRUE,
            payment_status='not_applicable'
    """, (MARKETER_USERNAME, MARKETER_PHONE, hash_password(MARKETER_PASSWORD)))

def current_month_label():
    return datetime.utcnow().strftime("%Y-%m")

def count_months_inclusive(start_month, end_month):
    try:
        start = datetime.strptime(str(start_month), "%Y-%m")
        end = datetime.strptime(str(end_month), "%Y-%m")
    except (TypeError, ValueError):
        return 0
    if end < start:
        return 0
    return (end.year - start.year) * 12 + end.month - start.month + 1

def has_role(required_role):
    current_role = session.get("role")
    if required_role is None:
        return True
    if isinstance(required_role, (set, tuple, list)):
        return current_role in required_role
    return current_role == required_role

def login_required(role=None):
    def deco(fn):
        @wraps(fn)
        def wrap(*a, **kw):
            if "user_id" not in session:
                return redirect(url_for("login"))
            if role and not has_role(role):
                flash("Vous nâ€™avez pas lâ€™autorisation.", "danger")
                return redirect(url_for("home"))
            return fn(*a, **kw)
        return wrap
    return deco

def admin_login_required(fn):
    return login_required(ADMIN_ROLES)(fn)

def developer_required(fn):
    return login_required("developer")(fn)

def finance_login_required(fn):
    return login_required(FINANCE_ROLES)(fn)

def marketer_login_required(fn):
    return login_required(MARKETER_ROLES)(fn)

def is_developer():
    return session.get("role") == "developer"

def target_is_developer(uid):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT role FROM users WHERE id=%s", (uid,))
        row = cur.fetchone()
        cur.close()
    return bool(row and row[0] == "developer")

def target_is_protected(uid):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT role FROM users WHERE id=%s", (uid,))
        row = cur.fetchone()
        cur.close()
    return bool(row and row[0] in {"developer", "finance", "marketer"})

# ===== Routes =====
@app.route("/")
def home():
    if "user_id" in session:
        r = session["role"]
        return redirect(url_for("admin_dashboard" if r in ADMIN_ROLES else
                                "finance_dashboard" if r=="finance" else
                                "offers_dashboard" if r=="marketer" else
                                "teacher_dashboard" if r=="teacher" else
                                "student_dashboard"))
    return render_template("public_results_home.html")

@app.route("/results")
def public_results_home():
    return render_template("public_results_home.html")

@app.route("/results/<exam_type>")
def public_results_search(exam_type):
    if exam_type not in RESULT_EXAM_TYPES:
        return redirect(url_for("home"))
    return render_template(
        "public_results_search.html",
        exam_type=exam_type,
        title=RESULT_EXAM_LABELS.get(exam_type, "نتائج المسابقات"),
    )

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/privacy")
def privacy():
    return render_template("privacy.html")

@app.route("/contact")
def contact():
    return render_template("contact.html")

@app.route("/robots.txt")
def robots_txt():
    sitemap_url = url_for("sitemap_xml", _external=True)
    body = "\n".join([
        "User-agent: *",
        "Allow: /",
        "Disallow: /admin",
        "Disallow: /finance",
        "Disallow: /teacher",
        "Disallow: /student",
        "Disallow: /payment",
        "Disallow: /verify",
        "Disallow: /reset-verify",
        "Disallow: /forgot",
        "Disallow: /logout",
        "Disallow: /onboarding",
        f"Sitemap: {sitemap_url}",
        "",
    ])
    return Response(body, mimetype="text/plain")

@app.route("/sitemap.xml")
def sitemap_xml():
    pages = [
        ("home", "1.0", "daily"),
        ("archive", "0.8", "daily"),
        ("about", "0.7", "monthly"),
        ("contact", "0.6", "monthly"),
        ("privacy", "0.3", "yearly"),
    ]
    lastmod = datetime.utcnow().strftime("%Y-%m-%d")
    urls = [
        f"""  <url>
    <loc>{url_for(endpoint, _external=True)}</loc>
    <lastmod>{lastmod}</lastmod>
    <changefreq>{changefreq}</changefreq>
    <priority>{priority}</priority>
  </url>"""
        for endpoint, priority, changefreq in pages
    ]
    body = """<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
%s
</urlset>
""" % "\n".join(urls)
    return Response(body, mimetype="application/xml")

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "GET":
        if "user_id" in session:
            return redirect(url_for("home"))
        return render_template("login.html")

    identifier = (request.form.get("phone") or request.form.get("username") or "").strip()
    password = request.form.get("password","")
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT * FROM users
                       WHERE (username=%s OR phone=%s) AND password=%s
                       LIMIT 1""",
                    (identifier, identifier, hash_password(password)))
        u = cur.fetchone(); cur.close()
    if not u:
        flash("رقم الهاتف أو كلمة المرور غير صحيحة.", "danger"); return redirect(url_for("login"))
    if u["phone_verified"] != 1:
        code = create_otp(u["phone"], "register")
        send_sms(u["phone"], f"Votre code de vÃ©rification est : {code}")
        session["pending_phone"] = u["phone"]
        flash("يجب التحقق من رقم الهاتف أولاً.", "warning")
        return redirect(url_for("verify_phone"))
    if u["status"] != "active":
        flash("الحساب بانتظار موافقة الإدارة.", "warning"); return redirect(url_for("login"))
    session.update(user_id=u["id"], role=u["role"], level=u["level"], subject=u.get("subject"))
    return redirect(url_for("home"))

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("home"))

@app.route("/change-password", methods=["GET", "POST"])
@login_required()
def change_password():
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not current_password or not new_password or not confirm_password:
            flash("Tous les champs sont requis.", "danger")
            return redirect(url_for("change_password"))
        if new_password != confirm_password:
            flash("Le nouveau mot de passe et sa confirmation ne correspondent pas.", "danger")
            return redirect(url_for("change_password"))
        if len(new_password) < 4:
            flash("Le nouveau mot de passe doit contenir au moins 4 caractères.", "danger")
            return redirect(url_for("change_password"))

        with db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT password FROM users WHERE id=%s", (session["user_id"],))
            row = cur.fetchone()
            if not row or row[0] != hash_password(current_password):
                cur.close()
                flash("Mot de passe actuel incorrect.", "danger")
                return redirect(url_for("change_password"))
            cur.execute(
                "UPDATE users SET password=%s WHERE id=%s",
                (hash_password(new_password), session["user_id"]),
            )
            conn.commit()
            cur.close()

        flash("Mot de passe modifié.", "success")
        return redirect(url_for("home"))

    return render_template("change_password.html")

# --- Inscription Ã©tudiant: donnÃ©es de base Ø«Ù… onboarding Ø«Ù… paiement
@app.route("/register", methods=["GET","POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username","").strip()
        phone    = request.form.get("phone","").strip()
        password = request.form.get("password","")
        confirm  = request.form.get("confirm_password","")

        if not username or not phone or not password or not confirm:
            flash("Veuillez complÃ©ter tous les champs.", "danger"); return redirect(url_for("register"))
        if password != confirm:
            flash("Le code et sa confirmation ne correspondent pas.", "danger"); return redirect(url_for("register"))

        session["pending_registration"] = {
            "username": username,
            "phone": phone,
            "password": password,
        }
        return redirect(url_for("courses"))
    return render_template("register.html")

COURSES = [
    {
        "code": "MPI",
        "title": "MPI",
        "subtitle": "Ø§Ù„Ø±ÙŠØ§Ø¶ÙŠØ§ØªØŒ Ø§Ù„ÙÙŠØ²ÙŠØ§Ø¡ ÙˆØ§Ù„Ø¥Ø¹Ù„Ø§Ù…ÙŠØ©",
        "description": "Ø¯Ø±ÙˆØ³ Ù…Ø±ÙƒØ²Ø© ÙˆØªÙ…Ø§Ø±ÙŠÙ† Ù…Ø­Ù„ÙˆÙ„Ø© Ù„Ù„ØªØ­Ø¶ÙŠØ± Ø¨Ø«Ù‚Ø©.",
        "badge": "Ù…ØªØ§Ø­",
        "icon": "ðŸ§®",
        "theme": "blue",
    },
    {
        "code": "PC",
        "title": "PC",
        "subtitle": "Ø§Ù„ÙÙŠØ²ÙŠØ§Ø¡ ÙˆØ§Ù„ÙƒÙŠÙ…ÙŠØ§Ø¡",
        "description": "Ø´Ø±Ø­ Ù…Ø¨Ø³Ø· Ù„Ù„ØªØ¬Ø§Ø±Ø¨ ÙˆØ§Ù„Ù‚ÙˆØ§Ù†ÙŠÙ† ÙˆØ§Ù„ØªÙ…Ø§Ø±ÙŠÙ†.",
        "badge": "Ù…ØªØ§Ø­",
        "icon": "âš—",
        "theme": "green",
    },
    {
        "code": "BG",
        "title": "BG",
        "subtitle": "Ø¹Ù„ÙˆÙ… Ø§Ù„Ø­ÙŠØ§Ø© ÙˆØ§Ù„Ø£Ø±Ø¶",
        "description": "Ù…Ù„Ø®ØµØ§Øª Ù…Ù†Ø¸Ù…Ø© ÙˆØ±Ø³ÙˆÙ… ØªÙˆØ¶ÙŠØ­ÙŠØ© Ù„Ù„Ù…Ø±Ø§Ø¬Ø¹Ø©.",
        "badge": "Ù…ØªØ§Ø­",
        "icon": "ðŸ§¬",
        "theme": "purple",
    },
    {
        "code": "BAC",
        "title": "BAC",
        "subtitle": "Ø¨Ø§ÙƒØ§Ù„ÙˆØ±ÙŠØ§",
        "description": "Ø¨Ø±Ù†Ø§Ù…Ø¬ Ù…Ø±Ø§Ø¬Ø¹Ø© Ø´Ø§Ù…Ù„ Ù„Ø§Ø¬ØªÙŠØ§Ø² Ø§Ù„Ø§Ù…ØªØ­Ø§Ù†.",
        "badge": "Ù‚Ø±ÙŠØ¨Ø§",
        "icon": "ðŸŽ“",
        "theme": "orange",
    },
    {
        "code": "BREVET",
        "title": "BREVET",
        "subtitle": "Ø´Ù‡Ø§Ø¯Ø© Ø®ØªÙ… Ø§Ù„Ø¯Ø±ÙˆØ³ Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯ÙŠØ©",
        "description": "Ø¯Ø±ÙˆØ³ ÙˆØªØ·Ø¨ÙŠÙ‚Ø§Øª Ø­Ø³Ø¨ Ø§Ù„Ø¨Ø±Ù†Ø§Ù…Ø¬.",
        "badge": "Ù‚Ø±ÙŠØ¨Ø§",
        "icon": "ðŸ“˜",
        "theme": "cyan",
    },
    {
        "code": "DESIGN",
        "title": "ØªØµÙ…ÙŠÙ… Ø¬Ø±Ø§ÙÙŠÙƒ",
        "subtitle": "ØªØ¹Ù„Ù… Ø§Ù„ØªØµÙ…ÙŠÙ… Ù…Ù† Ø§Ù„ØµÙØ±",
        "description": "Ø£Ø³Ø§Ø³ÙŠØ§Øª Ø§Ù„ØªØµÙ…ÙŠÙ… ÙˆØ£Ø¯ÙˆØ§Øª Ø§Ù„Ø¹Ù…Ù„.",
        "badge": "Ù‚Ø±ÙŠØ¨Ø§",
        "icon": "ðŸŽ¨",
        "theme": "pink",
    },
    {
        "code": "AI",
        "title": "Ø°ÙƒØ§Ø¡ Ø§ØµØ·Ù†Ø§Ø¹ÙŠ",
        "subtitle": "Ù…Ø¨Ø§Ø¯Ø¦ Ø§Ù„Ø°ÙƒØ§Ø¡ Ø§Ù„Ø§ØµØ·Ù†Ø§Ø¹ÙŠ",
        "description": "ØªØ¹Ù„Ù… Ø§Ù„Ù…ÙØ§Ù‡ÙŠÙ… Ø§Ù„Ø£Ø³Ø§Ø³ÙŠØ© ÙˆØ§Ù„ØªØ·Ø¨ÙŠÙ‚Ø§Øª.",
        "badge": "Ù‚Ø±ÙŠØ¨Ø§",
        "icon": "ðŸ¤–",
        "theme": "indigo",
    },
]

def ensure_courses_table():
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(100) NOT NULL UNIQUE,
                phone VARCHAR(30) NOT NULL UNIQUE,
                password VARCHAR(64) NOT NULL,
                role VARCHAR(20) NOT NULL DEFAULT 'student',
                level VARCHAR(40),
                subject VARCHAR(80),
                status VARCHAR(20) NOT NULL DEFAULT 'pending',
                phone_verified BOOLEAN NOT NULL DEFAULT FALSE,
                payment_image VARCHAR(255),
                payment_status VARCHAR(20),
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS verification_codes (
                id SERIAL PRIMARY KEY,
                phone VARCHAR(30) NOT NULL,
                code VARCHAR(10) NOT NULL,
                purpose VARCHAR(30) NOT NULL,
                used BOOLEAN NOT NULL DEFAULT FALSE,
                expires_at TIMESTAMP NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_verification_codes_lookup
            ON verification_codes (phone, purpose, code, used, expires_at)
        """)
        cur.execute("SELECT to_regclass('public.courses')")
        table_exists = cur.fetchone()[0] is not None
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS subject VARCHAR(80)")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS selected_subjects TEXT")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS payment_sender_phone VARCHAR(30)")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS payment_status VARCHAR(20)")
        cur.execute("""
            UPDATE users
            SET payment_status = CASE
                WHEN role <> 'student' THEN 'not_applicable'
                WHEN payment_image IS NOT NULL OR payment_sender_phone IS NOT NULL THEN 'pending'
                ELSE 'unpaid'
            END
            WHERE payment_status IS NULL
        """)
        cur.execute("ALTER TABLE users ALTER COLUMN payment_status SET DEFAULT 'pending'")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS courses (
                id SERIAL PRIMARY KEY,
                code VARCHAR(40) NOT NULL UNIQUE,
                title VARCHAR(100) NOT NULL,
                subtitle VARCHAR(255) NOT NULL DEFAULT '',
                description TEXT,
                badge VARCHAR(40) DEFAULT '',
                icon VARCHAR(20) DEFAULT 'ðŸ“˜',
                theme VARCHAR(30) DEFAULT 'blue',
                sort_order INT DEFAULT 0,
                active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS course_subjects (
                id SERIAL PRIMARY KEY,
                course_code VARCHAR(40) NOT NULL REFERENCES courses(code) ON DELETE CASCADE,
                subject VARCHAR(80) NOT NULL,
                sort_order INT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE (course_code, subject)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lessons (
                id SERIAL PRIMARY KEY,
                subject VARCHAR(80) NOT NULL,
                chapter_title VARCHAR(255) NOT NULL,
                level VARCHAR(40) NOT NULL,
                video_file VARCHAR(255),
                pdf_file VARCHAR(255),
                video_url TEXT,
                uploaded_by INT REFERENCES users(id) ON DELETE SET NULL,
                uploaded_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS free_pdfs (
                id SERIAL PRIMARY KEY,
                course_code VARCHAR(40) NOT NULL REFERENCES courses(code) ON DELETE CASCADE,
                subject VARCHAR(80) NOT NULL,
                title VARCHAR(255) NOT NULL,
                drive_url TEXT NOT NULL,
                sort_order INT DEFAULT 0,
                active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS notifications (
                id SERIAL PRIMARY KEY,
                title VARCHAR(255) NOT NULL,
                body TEXT NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key VARCHAR(80) PRIMARY KEY,
                value TEXT NOT NULL DEFAULT '',
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS active_site_visitors (
                visitor_key VARCHAR(80) PRIMARY KEY,
                last_seen TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS offer_slides (
                id SERIAL PRIMARY KEY,
                title VARCHAR(160) NOT NULL DEFAULT '',
                image_file VARCHAR(255),
                image_data BYTEA,
                image_mime VARCHAR(80),
                duration_seconds INT NOT NULL DEFAULT 5,
                sort_order INT NOT NULL DEFAULT 0,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_by INT REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("ALTER TABLE offer_slides ALTER COLUMN image_file DROP NOT NULL")
        cur.execute("ALTER TABLE offer_slides ADD COLUMN IF NOT EXISTS image_data BYTEA")
        cur.execute("ALTER TABLE offer_slides ADD COLUMN IF NOT EXISTS image_mime VARCHAR(80)")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS in_person_students (
                id SERIAL PRIMARY KEY,
                full_name VARCHAR(150) NOT NULL,
                phone VARCHAR(30),
                course_name VARCHAR(100) NOT NULL DEFAULT '',
                monthly_amount NUMERIC(12,2) NOT NULL DEFAULT 0,
                start_month VARCHAR(7) NOT NULL,
                delivery_type VARCHAR(20) NOT NULL DEFAULT 'near',
                notes TEXT,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("ALTER TABLE in_person_students ADD COLUMN IF NOT EXISTS delivery_type VARCHAR(20) DEFAULT 'near'")
        cur.execute("UPDATE in_person_students SET delivery_type='near' WHERE delivery_type IS NULL")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS in_person_payments (
                id SERIAL PRIMARY KEY,
                student_id INT NOT NULL REFERENCES in_person_students(id) ON DELETE CASCADE,
                month_label VARCHAR(7) NOT NULL,
                amount NUMERIC(12,2) NOT NULL DEFAULT 0,
                paid_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                recorded_by INT REFERENCES users(id) ON DELETE SET NULL,
                UNIQUE (student_id, month_label)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS finance_categories (
                id SERIAL PRIMARY KEY,
                name VARCHAR(120) NOT NULL,
                category_type VARCHAR(20) NOT NULL,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE (name, category_type)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS result_uploads (
                id SERIAL PRIMARY KEY,
                exam_type VARCHAR(40) NOT NULL,
                original_filename VARCHAR(255) NOT NULL,
                rows_imported INT NOT NULL DEFAULT 0,
                uploaded_by INT REFERENCES users(id) ON DELETE SET NULL,
                uploaded_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS national_exam_results (
                id SERIAL PRIMARY KEY,
                exam_type VARCHAR(40) NOT NULL,
                candidate_number VARCHAR(80),
                full_name VARCHAR(255) NOT NULL,
                birth_place VARCHAR(160),
                birth_date VARCHAR(80),
                wilaya VARCHAR(160),
                moughataa VARCHAR(160),
                center_name VARCHAR(255),
                score VARCHAR(80),
                decision VARCHAR(160),
                rank VARCHAR(80),
                raw_data JSONB NOT NULL DEFAULT '{}'::jsonb,
                upload_id INT REFERENCES result_uploads(id) ON DELETE SET NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("ALTER TABLE national_exam_results ADD COLUMN IF NOT EXISTS rank VARCHAR(80)")
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_lessons_level_subject
            ON lessons (level, subject, uploaded_at DESC)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_lessons_uploaded_by
            ON lessons (uploaded_by, uploaded_at DESC)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_free_pdfs_course_subject
            ON free_pdfs (course_code, subject, sort_order ASC, id DESC)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_notifications_created_at
            ON notifications (created_at DESC)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_offer_slides_active_order
            ON offer_slides (active, sort_order ASC, id DESC)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_active_site_visitors_last_seen
            ON active_site_visitors (last_seen DESC)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_in_person_students_active
            ON in_person_students (active, created_at DESC)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_in_person_payments_student_month
            ON in_person_payments (student_id, month_label)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_finance_categories_active_type
            ON finance_categories (active, category_type, name)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_national_results_exam_number
            ON national_exam_results (exam_type, candidate_number)
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_national_results_exam_name
            ON national_exam_results (exam_type, lower(full_name))
        """)
        if not table_exists:
            for index, course in enumerate(COURSES, start=1):
                cur.execute("""
                    INSERT INTO courses
                        (code, title, subtitle, description, badge, icon, theme, sort_order, active)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
                """, (
                    course["code"], course["title"], course["subtitle"], course["description"],
                    course["badge"], course["icon"], course["theme"], index
                ))
        cur.execute("SELECT COUNT(*) FROM course_subjects")
        if cur.fetchone()[0] == 0:
            default_subjects = ["Math", "Physique", "Chimie", "Science naturelle"]
            cur.execute("SELECT code FROM courses ORDER BY sort_order ASC, id ASC")
            existing_courses = [row[0] for row in cur.fetchall()]
            for course_code in existing_courses:
                for index, subject in enumerate(default_subjects, start=1):
                    cur.execute("""
                        INSERT INTO course_subjects (course_code, subject, sort_order)
                        VALUES (%s,%s,%s)
                        ON CONFLICT (course_code, subject) DO NOTHING
                    """, (course_code, subject, index))
        ensure_default_finance_user(cur)
        ensure_default_marketer_user(cur)
        conn.commit()
        cur.close()

def fetch_app_settings():
    defaults = {
        "paymentNumber": os.getenv("PAYMENT_NUMBER", "22334455"),
        "paymentAmount": os.getenv("PAYMENT_AMOUNT", "غير محدد"),
        "offerTextTitle": "",
        "offerTextBody": "",
        "offerTextActive": "true",
        "visitorCountOffset": "0",
        "visitorAutoAddAmount": "0",
        "visitorAutoAddEverySeconds": "60",
        "visitorAutoSubtractAmount": "0",
        "visitorAutoSubtractEverySeconds": "60",
        "visitorAutomationStartedAt": "0",
    }
    try:
        ensure_courses_table()
        with db() as conn:
            cur = dict_cursor(conn)
            cur.execute("SELECT key, value FROM app_settings")
            rows = cur.fetchall()
            cur.close()
        settings = defaults.copy()
        settings.update({row["key"]: row["value"] for row in rows})
        return settings
    except psycopg2.OperationalError as e:
        print("Database unavailable, using default settings:", e)
        return defaults

def save_app_settings(values):
    ensure_courses_table()
    allowed_keys = {
        "paymentNumber",
        "paymentAmount",
        "offerTextTitle",
        "offerTextBody",
        "offerTextActive",
        "visitorCountOffset",
        "visitorAutoAddAmount",
        "visitorAutoAddEverySeconds",
        "visitorAutoSubtractAmount",
        "visitorAutoSubtractEverySeconds",
        "visitorAutomationStartedAt",
    }
    cleaned = {
        key: str(value).strip()
        for key, value in values.items()
        if key in allowed_keys and value is not None
    }
    if not cleaned:
        return fetch_app_settings()
    with db() as conn:
        cur = conn.cursor()
        for key, value in cleaned.items():
            cur.execute("""
                INSERT INTO app_settings (key, value, updated_at)
                VALUES (%s,%s,CURRENT_TIMESTAMP)
                ON CONFLICT (key)
                DO UPDATE SET value=EXCLUDED.value, updated_at=CURRENT_TIMESTAMP
            """, (key, value))
        conn.commit()
        cur.close()
    return fetch_app_settings()

def offer_slide_payload(row):
    return {
        "id": str(row["id"]),
        "title": row.get("title") or "",
        "imageUrl": url_for(
            "api_offer_image",
            slide_id=row["id"],
            _external=True,
        ),
        "durationSeconds": int(row.get("duration_seconds") or 5),
        "sortOrder": int(row.get("sort_order") or 0),
        "active": bool(row.get("active")),
    }

def fetch_offer_slides(active_only=True):
    ensure_courses_table()
    with db() as conn:
        cur = dict_cursor(conn)
        where = "WHERE active=TRUE" if active_only else ""
        cur.execute(f"""
            SELECT *
            FROM offer_slides
            {where}
            ORDER BY sort_order ASC, id DESC
        """)
        rows = cur.fetchall()
        cur.close()
    return rows

def fetch_courses(active_only=True):
    try:
        ensure_courses_table()
        with db() as conn:
            cur = dict_cursor(conn)
            where = "WHERE active=TRUE" if active_only else ""
            cur.execute(f"""SELECT id, code, title, subtitle, description, badge, icon, theme, sort_order, active
                            FROM courses {where}
                            ORDER BY sort_order ASC, id ASC""")
            rows = cur.fetchall()
            cur.close()
        return rows
    except psycopg2.OperationalError as e:
        print("Database unavailable, using default courses:", e)
        return [dict(course, id=index, sort_order=index, active=True)
                for index, course in enumerate(COURSES, start=1)]

def fetch_course_subjects(course_code=None):
    try:
        ensure_courses_table()
        with db() as conn:
            cur = dict_cursor(conn)
            if course_code:
                cur.execute("""SELECT id, course_code, subject, sort_order
                               FROM course_subjects
                               WHERE course_code=%s
                               ORDER BY sort_order ASC, id ASC""", (course_code,))
            else:
                cur.execute("""SELECT id, course_code, subject, sort_order
                               FROM course_subjects
                               ORDER BY course_code ASC, sort_order ASC, id ASC""")
            rows = cur.fetchall()
            cur.close()
        return rows
    except psycopg2.OperationalError as e:
        print("Database unavailable, using default subjects:", e)
        default_subjects = ["Math", "Physique", "Chimie", "Science naturelle"]
        courses = [course_code] if course_code else [course["code"] for course in COURSES]
        return [
            {"id": index, "course_code": code, "subject": subject, "sort_order": index}
            for code in courses
            for index, subject in enumerate(default_subjects, start=1)
        ]

def fetch_free_pdfs(active_only=True, course_code=None, subject=None):
    try:
        ensure_courses_table()
        with db() as conn:
            cur = dict_cursor(conn)
            clauses = []
            params = []
            if active_only:
                clauses.append("fp.active=TRUE")
            if course_code:
                clauses.append("fp.course_code=%s")
                params.append(course_code)
            if subject:
                clauses.append("fp.subject=%s")
                params.append(subject)
            where = "WHERE " + " AND ".join(clauses) if clauses else ""
            cur.execute(f"""
                SELECT fp.id, fp.course_code, c.title AS course_title, fp.subject,
                       fp.title, fp.drive_url, fp.sort_order, fp.active, fp.created_at
                FROM free_pdfs fp
                JOIN courses c ON c.code = fp.course_code
                {where}
                ORDER BY c.sort_order ASC, c.id ASC,
                         fp.subject ASC, fp.sort_order ASC, fp.id DESC
            """, params)
            rows = cur.fetchall()
            cur.close()
    except psycopg2.OperationalError as e:
        print("Database unavailable, no free PDFs loaded:", e)
        return []

    grouped = []
    by_course = {}
    by_subject = {}
    for row in rows:
        row["preview_url"] = google_drive_preview_url(row["drive_url"])
        course = by_course.get(row["course_code"])
        if not course:
            course = {
                "level": row["course_code"],
                "title": row["course_title"],
                "subjects": []
            }
            by_course[row["course_code"]] = course
            grouped.append(course)
        subject_key = (row["course_code"], row["subject"])
        subject = by_subject.get(subject_key)
        if not subject:
            subject = {"name": row["subject"], "pdfs": []}
            by_subject[subject_key] = subject
            course["subjects"].append(subject)
        subject["pdfs"].append(row)
    return grouped

_schema_ready = False
_active_visitors = {}
_visitor_settings_cache = {"loaded_at": 0, "settings": {}}

@app.before_request
def ensure_schema_ready():
    global _schema_ready
    if request.endpoint in {"robots_txt", "sitemap_xml", "static"}:
        return
    if not _schema_ready:
        try:
            ensure_courses_table()
            _schema_ready = True
        except psycopg2.OperationalError as e:
            print("Database unavailable during schema check:", e)

@app.route("/courses", methods=["GET","POST"])
def courses():
    pending = session.get("pending_registration")
    if not pending:
        return redirect(url_for("register"))
    available_courses = fetch_courses(active_only=True)
    if request.method == "POST":
        selected = request.form.get("course")
        course = next((c for c in available_courses if c["code"] == selected), None)
        if not course:
            flash("Veuillez choisir une formation.", "danger")
            return redirect(url_for("courses"))
        pending["level"] = course["code"]
        pending["course_title"] = course["title"]
        session["pending_registration"] = pending
        return redirect(url_for("onboarding"))
    return render_template("courses.html", courses=available_courses)

@app.route("/archive")
def archive():
    courses = fetch_courses(active_only=True)
    selected_level = request.args.get("level","").strip()
    selected_subject = request.args.get("subject","").strip()
    valid_levels = {course["code"] for course in courses}
    if selected_level and selected_level not in valid_levels:
        selected_level = ""
        selected_subject = ""

    subjects = fetch_course_subjects(selected_level) if selected_level else []
    valid_subjects = {row["subject"] for row in subjects}
    if selected_subject and selected_subject not in valid_subjects:
        selected_subject = ""

    free_pdfs = []
    if selected_level and selected_subject:
        free_pdfs = fetch_free_pdfs(active_only=True,
                                    course_code=selected_level,
                                    subject=selected_subject)
    return render_template("archive.html", free_pdfs=free_pdfs, courses=courses,
                           subjects=subjects, selected_level=selected_level,
                           selected_subject=selected_subject)

@app.route("/onboarding")
def onboarding():
    if "pending_registration" not in session:
        return redirect(url_for("register"))
    return render_template("onboarding.html")

@app.route("/payment", methods=["GET","POST"])
def payment():
    pending = session.get("pending_registration")
    if not pending:
        return redirect(url_for("register"))
    if request.method == "POST":
        pfile = request.files.get("payment_image")
        if not pfile or pfile.filename == "":
            flash("Merci de tÃ©lÃ©charger lâ€™attestation de paiement.", "warning")
            return redirect(url_for("payment"))
        if not allowed(pfile.filename, IMG_EXT):
            flash("Fichier non autorisÃ© (jpg/png/webp/pdf).", "danger")
            return redirect(url_for("payment"))

        username = pending["username"]
        phone = pending["phone"]
        password = pending["password"]
        level = pending.get("level", "4as")

        base = secure_filename(pfile.filename)
        fname = f"{username}_{int(time.time())}_{base}"
        pfile.save(os.path.join(PAY_DIR, fname))

        try:
            with db() as conn:
                cur = conn.cursor()
                cur.execute("""INSERT INTO users
                               (username, phone, password, role, level, status, phone_verified, payment_image, payment_status)
                               VALUES (%s,%s,%s,'student',%s,'pending',TRUE,%s,'pending')""",
                            (username, phone, hash_password(password), level, fname))
                conn.commit()
                cur.close()
            session.pop("pending_registration", None)
            flash("Votre compte a Ã©tÃ© crÃ©Ã©. Il attend lâ€™approbation de lâ€™administrateur.", "success")
            return redirect(url_for("home"))
        except IntegrityError as e:
            msg = "Nom dâ€™utilisateur dÃ©jÃ  utilisÃ©."
            if "phone" in str(e):
                msg = "NumÃ©ro de tÃ©lÃ©phone dÃ©jÃ  utilisÃ©."
            flash(msg, "danger")
            return redirect(url_for("payment"))
    return render_template("payment.html")

# --- Saisie du code OTP (vÃ©rification du tÃ©lÃ©phone)
@app.route("/verify", methods=["GET","POST"])
def verify_phone():
    phone = session.get("pending_phone","")
    if request.method == "POST":
        code = request.form.get("otp","").strip()
        phone_form = request.form.get("phone","").strip()
        if phone_form: phone = phone_form
        if not phone or not code:
            flash("Le numÃ©ro et le code sont requis.", "danger"); return redirect(url_for("verify_phone"))
        if verify_otp(phone, "register", code):
            with db() as conn:
                cur = conn.cursor()
                cur.execute("UPDATE users SET phone_verified=TRUE WHERE phone=%s", (phone,))
                conn.commit(); cur.close()
            session.pop("pending_phone", None)
            flash("NumÃ©ro vÃ©rifiÃ©. Votre compte attend lâ€™approbation de lâ€™administrateur.", "success")
            return redirect(url_for("home"))
        else:
            flash("Code invalide ou expirÃ©.", "danger")
    return render_template("verify.html", phone=phone)

# --- Mot de passe oubliÃ©
@app.route("/forgot", methods=["GET","POST"])
def forgot():
    if request.method == "POST":
        phone = request.form.get("phone","").strip()
        with db() as conn:
            cur = dict_cursor(conn)
            cur.execute("SELECT id FROM users WHERE phone=%s", (phone,))
            u = cur.fetchone(); cur.close()
        if not u:
            flash("Aucun compte liÃ© Ã  ce numÃ©ro.", "danger"); return redirect(url_for("forgot"))
        code = create_otp(phone, "reset")
        send_sms(phone, f"Code de rÃ©initialisation : {code}")
        session["reset_phone"] = phone
        flash("Un code de rÃ©initialisation a Ã©tÃ© envoyÃ© par SMS.", "success")
        return redirect(url_for("reset_verify"))
    return render_template("forgot.html")

@app.route("/reset-verify", methods=["GET","POST"])
def reset_verify():
    phone = session.get("reset_phone","")
    if request.method == "POST":
        code = request.form.get("code","").strip()
        newp = request.form.get("new_password","")
        phone_form = request.form.get("phone","").strip()
        if phone_form: phone = phone_form
        if not phone or not code or not newp:
            flash("Tous les champs sont requis.", "danger"); return redirect(url_for("reset_verify"))
        if verify_otp(phone, "reset", code):
            with db() as conn:
                cur = conn.cursor()
                cur.execute("UPDATE users SET password=%s WHERE phone=%s", (hash_password(newp), phone))
                conn.commit(); cur.close()
            session.pop("reset_phone", None)
            flash("Mot de passe rÃ©initialisÃ©. Vous pouvez vous connecter.", "success")
            return redirect(url_for("home"))
        else:
            flash("Code invalide ou expirÃ©.", "danger")
    return render_template("reset_verify.html", phone=phone)

# ===== Offers / Marketer =====
@app.route("/offers")
@marketer_login_required
def offers_dashboard():
    slides = fetch_offer_slides(active_only=False)
    return render_template(
        "offers.html",
        slides=slides,
        settings=fetch_app_settings(),
        is_developer=is_developer(),
    )

@app.post("/offers/text")
@marketer_login_required
def offers_text_update():
    save_app_settings({
        "offerTextTitle": request.form.get("offer_text_title", "").strip(),
        "offerTextBody": request.form.get("offer_text_body", "").strip(),
        "offerTextActive": "true" if request.form.get("offer_text_active") == "on" else "false",
    })
    flash("تم حفظ نص العروض.", "success")
    return redirect(url_for("offers_dashboard"))

@app.post("/admin/visitor-offset")
@developer_required
def developer_update_visitor_offset():
    def positive_int(field, default="0", minimum=0):
        value = request.form.get(field, default).strip()
        number = int(value or default)
        return max(minimum, number)

    def interval_seconds(value_field, unit_field):
        value = positive_int(value_field, "1", minimum=1)
        unit = request.form.get(unit_field, "seconds")
        return value * 60 if unit == "minutes" else value

    try:
        offset = positive_int("visitor_count_offset")
        add_amount = positive_int("visitor_auto_add_amount")
        add_every = interval_seconds("visitor_auto_add_every", "visitor_auto_add_unit")
        subtract_amount = positive_int("visitor_auto_subtract_amount")
        subtract_every = interval_seconds("visitor_auto_subtract_every", "visitor_auto_subtract_unit")
    except ValueError:
        flash("اكتب أرقاما صحيحة لإعدادات عداد الزوار.", "danger")
        return redirect(url_for("admin_dashboard"))

    save_app_settings({
        "visitorCountOffset": str(offset),
        "visitorAutoAddAmount": str(add_amount),
        "visitorAutoAddEverySeconds": str(add_every),
        "visitorAutoSubtractAmount": str(subtract_amount),
        "visitorAutoSubtractEverySeconds": str(subtract_every),
        "visitorAutomationStartedAt": str(int(time.time())),
    })
    flash("تم تحديث زيادة عداد الزوار.", "success")
    return redirect(url_for("admin_dashboard"))

@app.post("/offers/upload")
@marketer_login_required
def offers_upload():
    title = request.form.get("title", "").strip()
    duration_text = request.form.get("duration_seconds", "5").strip()
    sort_order_text = request.form.get("sort_order", "0").strip()
    image = request.files.get("image")
    if not image or image.filename == "":
        flash("اختر صورة العرض.", "danger")
        return redirect(url_for("offers_dashboard"))
    if not allowed(image.filename, OFFER_IMG_EXT):
        flash("الصيغ المدعومة: jpg / png / webp فقط.", "danger")
        return redirect(url_for("offers_dashboard"))
    try:
        duration_seconds = max(1, min(120, int(duration_text)))
    except ValueError:
        duration_seconds = 5
    try:
        sort_order = int(sort_order_text)
    except ValueError:
        sort_order = 0

    base = secure_filename(image.filename) or "offer.webp"
    filename = f"offer_{int(time.time())}_{base}"
    image_data = image.read()
    image_mime = image.mimetype or "application/octet-stream"
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO offer_slides
                (title, image_file, image_data, image_mime, duration_seconds, sort_order, active, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,TRUE,(SELECT id FROM users WHERE id=%s))
        """, (
            title,
            filename,
            psycopg2.Binary(image_data),
            image_mime,
            duration_seconds,
            sort_order,
            session.get("user_id"),
        ))
        conn.commit()
        cur.close()
    flash("تمت إضافة العرض.", "success")
    return redirect(url_for("offers_dashboard"))

@app.post("/offers/<int:slide_id>/update")
@marketer_login_required
def offers_update(slide_id):
    title = request.form.get("title", "").strip()
    active = request.form.get("active") == "on"
    try:
        duration_seconds = max(1, min(120, int(request.form.get("duration_seconds", "5"))))
    except ValueError:
        duration_seconds = 5
    try:
        sort_order = int(request.form.get("sort_order", "0"))
    except ValueError:
        sort_order = 0
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            UPDATE offer_slides
            SET title=%s, duration_seconds=%s, sort_order=%s, active=%s
            WHERE id=%s
        """, (title, duration_seconds, sort_order, active, slide_id))
        conn.commit()
        cur.close()
    flash("تم تحديث العرض.", "success")
    return redirect(url_for("offers_dashboard"))

@app.post("/offers/<int:slide_id>/delete")
@marketer_login_required
def offers_delete(slide_id):
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT image_file FROM offer_slides WHERE id=%s", (slide_id,))
        slide = cur.fetchone()
        cur.execute("DELETE FROM offer_slides WHERE id=%s", (slide_id,))
        conn.commit()
        cur.close()
    if slide:
        remove_upload(OFFERS_DIR, slide.get("image_file"))
    flash("تم حذف العرض.", "success")
    return redirect(url_for("offers_dashboard"))

@app.get("/api/offers/images/<int:slide_id>")
def api_offer_image(slide_id):
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT image_data, image_mime, image_file
            FROM offer_slides
            WHERE id=%s
        """, (slide_id,))
        slide = cur.fetchone()
        cur.close()
    if not slide:
        return "Not found", 404
    if slide.get("image_data") is not None:
        data = bytes(slide["image_data"])
        return Response(
            data,
            mimetype=slide.get("image_mime") or "application/octet-stream",
            headers={"Cache-Control": "public, max-age=86400"},
        )
    image_file = slide.get("image_file")
    if image_file:
        path = os.path.abspath(os.path.join(OFFERS_DIR, image_file))
        if os.path.exists(path):
            return redirect(url_for("static", filename=f"uploads/offers/{image_file}"))
    return "Not found", 404

# ===== Tableau de bord Admin =====
@app.route("/admin")
@admin_login_required
def admin_dashboard():
    with db() as conn:
        cur = dict_cursor(conn)
        if is_developer():
            cur.execute("""SELECT id,username,phone,role,level,subject,status,phone_verified,payment_image,payment_status
                           FROM users ORDER BY id DESC""")
        else:
            cur.execute("""SELECT id,username,phone,role,level,subject,status,phone_verified,payment_image,payment_status
                           FROM users WHERE role NOT IN ('developer','finance','marketer') ORDER BY id DESC""")
        users = cur.fetchall()
        cur.execute("""SELECT l.*, u.username AS uploader_name
                       FROM lessons l
                       LEFT JOIN users u ON u.id = l.uploaded_by
                       ORDER BY l.uploaded_at DESC, l.id DESC""")
        lessons = cur.fetchall(); cur.close()
    courses = fetch_courses(active_only=False)
    course_subjects = fetch_course_subjects()
    free_pdfs = fetch_free_pdfs(active_only=False)
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""
            SELECT r.exam_type, COUNT(*) AS total, MAX(r.created_at) AS last_import,
                   MAX(r.upload_id) AS upload_id
            FROM national_exam_results r
            GROUP BY r.exam_type
        """)
        result_summary_rows = cur.fetchall()
        cur.execute("""
            SELECT ru.id, ru.exam_type, ru.original_filename, ru.rows_imported,
                   ru.uploaded_at, u.username AS uploaded_by_name,
                   COALESCE(COUNT(r.id), 0) AS active_rows
            FROM result_uploads ru
            LEFT JOIN users u ON u.id = ru.uploaded_by
            LEFT JOIN national_exam_results r ON r.upload_id = ru.id
            GROUP BY ru.id, u.username
            ORDER BY ru.uploaded_at DESC, ru.id DESC
            LIMIT 12
        """)
        result_uploads = cur.fetchall()
        cur.close()
    result_summary = {row["exam_type"]: row for row in result_summary_rows}
    result_query = (request.args.get("result_q") or "").strip()
    result_exam = request.args.get("result_exam", "bac-first").strip()
    if result_exam not in RESULT_EXAM_TYPES:
        result_exam = "bac-first"
    result_matches = []
    if result_query.isdigit() or len(result_query) >= 2:
        with db() as conn:
            cur = dict_cursor(conn)
            if result_query.isdigit():
                normalized_number = normalize_candidate_number(result_query)
                if result_exam == "concours":
                    cur.execute(f"""
                        SELECT *
                        FROM national_exam_results
                        WHERE exam_type=%s
                          AND {concours_number_match_sql()}
                        ORDER BY id ASC
                        LIMIT 1
                    """, (result_exam, normalized_number, normalized_number))
                else:
                    cur.execute("""
                        SELECT *
                        FROM national_exam_results
                        WHERE exam_type=%s
                          AND COALESCE(NULLIF(regexp_replace(candidate_number, '\\D', '', 'g'), ''), '0')::bigint = %s::bigint
                        ORDER BY id ASC
                        LIMIT 1
                    """, (result_exam, normalized_number))
            else:
                if result_exam == "concours":
                    search_concours_by_name(cur, result_query)
                else:
                    cur.execute("""
                        SELECT *
                        FROM national_exam_results
                        WHERE exam_type=%s AND full_name ILIKE %s
                        ORDER BY full_name ASC
                        LIMIT 1
                    """, (result_exam, f"%{result_query}%"))
            result_matches = cur.fetchall()
            cur.close()
    return render_template("admin.html", users=users, courses=courses,
                           course_subjects=course_subjects, free_pdfs=free_pdfs,
                           lessons=lessons,
                           result_exam_labels=RESULT_EXAM_LABELS,
                           result_summary=result_summary,
                           result_uploads=result_uploads,
                           result_query=result_query,
                           result_exam=result_exam,
                           result_matches=result_matches,
                           settings=fetch_app_settings(),
                           is_developer=is_developer())

@app.route("/admin/activate/<int:uid>")
@admin_login_required
def activate_user(uid):
    if not is_developer() and target_is_protected(uid):
        flash("Action non autorisée.", "danger")
        return redirect(url_for("admin_dashboard"))
    with db() as conn:
        cur = conn.cursor(); cur.execute("UPDATE users SET status='active' WHERE id=%s", (uid,))
        conn.commit(); cur.close()
    flash("Compte activÃ©.", "success"); return redirect(url_for("admin_dashboard"))

@app.route("/admin/delete/<int:uid>")
@admin_login_required
def delete_user(uid):
    if not is_developer() and target_is_protected(uid):
        flash("Action non autorisée.", "danger")
        return redirect(url_for("admin_dashboard"))
    if is_developer() and uid == session.get("user_id"):
        flash("Vous ne pouvez pas supprimer votre propre compte.", "danger")
        return redirect(url_for("admin_dashboard"))
    result = delete_user_with_teacher_lessons(uid)
    if result["lessons_deleted"]:
        flash(f"Compte supprimÃ© avec {result['lessons_deleted']} leÃ§on(s).", "warning")
    else:
        flash("Compte supprimÃ©.", "warning")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/results/upload", methods=["GET", "POST"])
@developer_required
def admin_upload_results():
    if request.method == "GET":
        return redirect(url_for("admin_dashboard") + "#national-results")
    try:
        exam_type = request.form.get("exam_type", "").strip()
        if exam_type not in RESULT_EXAM_TYPES:
            flash("نوع المسابقة غير صحيح.", "danger")
            return redirect(url_for("admin_dashboard") + "#national-results")
        file = request.files.get("file")
        if not file or not file.filename:
            flash("يرجى اختيار ملف Excel.", "danger")
            return redirect(url_for("admin_dashboard") + "#national-results")
        if not allowed(file.filename, EXCEL_EXT):
            flash("الملفات المدعومة هي xlsx و xlsm فقط. من Google Sheets اختر: Download ثم Microsoft Excel (.xlsx).", "danger")
            return redirect(url_for("admin_dashboard") + "#national-results")
        parsed_rows = parse_results_workbook(file)
        if not parsed_rows:
            flash("لم يتم العثور على نتائج داخل الملف.", "warning")
            return redirect(url_for("admin_dashboard") + "#national-results")
        filename = secure_filename(file.filename) or "results.xlsx"
        insert_national_results(exam_type, filename, parsed_rows, session.get("user_id"))
        flash(f"تم استيراد {len(parsed_rows)} نتيجة في {RESULT_EXAM_LABELS[exam_type]}. ستظهر مباشرة داخل التطبيق.", "success")
    except Exception as exc:
        app.logger.exception("National results import failed")
        flash(f"تعذر معالجة ملف Excel: {exc}", "danger")
        return redirect(url_for("admin_dashboard") + "#national-results")
    return redirect(url_for("admin_dashboard") + "#national-results")

@app.post("/admin/results/uploads/<int:upload_id>/delete")
@developer_required
def admin_delete_results_upload(upload_id):
    try:
        with db() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM national_exam_results WHERE upload_id=%s", (upload_id,))
            deleted_rows = cur.rowcount
            cur.execute("DELETE FROM result_uploads WHERE id=%s", (upload_id,))
            deleted_uploads = cur.rowcount
            conn.commit()
            cur.close()
        if deleted_uploads:
            flash(f"تم حذف دفعة النتائج مع {deleted_rows} نتيجة مرتبطة بها.", "warning")
        else:
            flash("دفعة النتائج غير موجودة.", "warning")
    except Exception as exc:
        app.logger.exception("National results delete failed")
        flash(f"تعذر حذف دفعة النتائج: {exc}", "danger")
    return redirect(url_for("admin_dashboard") + "#national-results")

# ===== Tableau de bord Finance =====
@app.route("/finance")
@finance_login_required
def finance_dashboard():
    view_mode = request.args.get("view", "dashboard").strip()
    if view_mode not in {"dashboard", "remote", "near", "categories"}:
        view_mode = "dashboard"
    remote_category = request.args.get("category", "").strip()
    near_category = request.args.get("near_category", "").strip()
    status_filter = request.args.get("status", "all").strip()
    valid_filters = {"all", "paid", "unpaid", "pending"}
    if status_filter not in valid_filters:
        status_filter = "all"
    selected_month = request.args.get("month", current_month_label()).strip() or current_month_label()
    if not re.match(r"^\d{4}-\d{2}$", selected_month):
        selected_month = current_month_label()

    with db() as conn:
        cur = dict_cursor(conn)
        students = []
        cur.execute("""
            SELECT
                COUNT(*) FILTER (WHERE active=TRUE) AS total,
                COUNT(*) FILTER (WHERE active=TRUE AND delivery_type='remote') AS remote,
                COUNT(*) FILTER (WHERE active=TRUE AND delivery_type='near') AS near
            FROM in_person_students
        """)
        stats = cur.fetchone()
        near_clauses = ["s.active=TRUE"]
        near_params = [selected_month]
        selected_delivery_type = "remote" if view_mode == "remote" else "near"
        near_clauses.append("s.delivery_type=%s")
        near_params.append(selected_delivery_type)
        selected_manual_category = remote_category if view_mode == "remote" else near_category
        if view_mode in {"near", "remote"} and selected_manual_category:
            near_clauses.append("s.course_name=%s")
            near_params.append(selected_manual_category)
        near_where = " AND ".join(near_clauses)
        cur.execute(f"""
            SELECT s.*,
                   COALESCE(paid_counts.paid_months, 0) AS paid_months,
                   month_payment.id AS current_payment_id,
                   month_payment.amount AS current_payment_amount,
                   month_payment.paid_at AS current_paid_at
            FROM in_person_students s
            LEFT JOIN (
                SELECT student_id, COUNT(*) AS paid_months
                FROM in_person_payments
                GROUP BY student_id
            ) paid_counts ON paid_counts.student_id = s.id
            LEFT JOIN in_person_payments month_payment
                   ON month_payment.student_id = s.id
                  AND month_payment.month_label = %s
            WHERE {near_where}
            ORDER BY s.created_at DESC, s.id DESC
        """, near_params)
        in_person_students = cur.fetchall()
        payment_clauses = []
        payment_params = []
        payment_clauses.append("s.delivery_type=%s")
        payment_params.append(selected_delivery_type)
        if view_mode in {"near", "remote"} and selected_manual_category:
            payment_clauses.append("s.course_name=%s")
            payment_params.append(selected_manual_category)
        payment_where = "WHERE " + " AND ".join(payment_clauses) if payment_clauses else ""
        cur.execute(f"""
            SELECT p.*, s.full_name
            FROM in_person_payments p
            JOIN in_person_students s ON s.id = p.student_id
            {payment_where}
            ORDER BY p.month_label DESC, p.paid_at DESC, p.id DESC
            LIMIT 40
        """, payment_params)
        in_person_payments = cur.fetchall()
        cur.execute("""
            SELECT id, name, category_type, created_at
            FROM finance_categories
            WHERE active=TRUE
            ORDER BY
                CASE category_type WHEN 'course' THEN 1 WHEN 'section' THEN 2 ELSE 3 END,
                name ASC
        """)
        finance_categories = cur.fetchall()
        cur.execute("""
            SELECT DISTINCT course_name AS name, 'course' AS category_type
            FROM in_person_students
            WHERE active=TRUE AND delivery_type='near' AND course_name IS NOT NULL AND course_name <> ''
            ORDER BY name
        """)
        near_student_categories = cur.fetchall()
        cur.execute("""
            SELECT DISTINCT course_name AS name, 'course' AS category_type
            FROM in_person_students
            WHERE active=TRUE AND delivery_type='remote' AND course_name IS NOT NULL AND course_name <> ''
            ORDER BY name
        """)
        remote_student_categories = cur.fetchall()
        cur.close()

    near_stats = {"total": len(in_person_students), "paid_this_month": 0, "debt_total": 0}
    for student in in_person_students:
        paid_months = int(student.get("paid_months") or 0)
        due_months = count_months_inclusive(student.get("start_month"), selected_month)
        unpaid_months = max(due_months - paid_months, 0)
        monthly_amount = float(student.get("monthly_amount") or 0)
        student["unpaid_months"] = unpaid_months
        student["total_due"] = unpaid_months * monthly_amount
        student["paid_current_month"] = bool(student.get("current_payment_id"))
        if student["paid_current_month"]:
            near_stats["paid_this_month"] += 1
        near_stats["debt_total"] += student["total_due"]
    stats["paid"] = near_stats["paid_this_month"]
    stats["unpaid"] = max((stats.get("total") or 0) - near_stats["paid_this_month"], 0)
    stats["pending"] = 0

    remote_categories = []
    seen_remote_categories = set()
    for source in (finance_categories, remote_student_categories):
        for item in source:
            key = (item["name"], item["category_type"])
            if key in seen_remote_categories:
                continue
            seen_remote_categories.add(key)
            remote_categories.append({
                "name": item["name"],
                "category_type": item["category_type"],
            })

    near_categories = []
    seen_near_categories = set()
    for source in (finance_categories, near_student_categories):
        for item in source:
            key = (item["name"], item["category_type"])
            if key in seen_near_categories:
                continue
            seen_near_categories.add(key)
            near_categories.append({
                "name": item["name"],
                "category_type": item["category_type"],
            })

    return render_template("finance.html", students=students, stats=stats,
                           status_filter=status_filter,
                           view_mode=view_mode,
                           remote_category=remote_category,
                           near_category=near_category,
                           remote_categories=remote_categories,
                           near_categories=near_categories,
                           selected_month=selected_month,
                           in_person_students=in_person_students,
                           in_person_payments=in_person_payments,
                           finance_categories=finance_categories,
                           near_stats=near_stats)

@app.post("/finance/users/<int:uid>/payment")
@finance_login_required
def finance_update_payment(uid):
    payment_status = request.form.get("payment_status", "").strip()
    if payment_status not in {"paid", "unpaid", "pending"}:
        flash("حالة الدفع غير صحيحة.", "danger")
        return redirect(url_for("finance_dashboard"))
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            UPDATE users
            SET payment_status=%s
            WHERE id=%s AND role='student'
        """, (payment_status, uid))
        updated = cur.rowcount
        conn.commit()
        cur.close()
    flash("تم تحديث حالة الدفع." if updated else "لم يتم العثور على الطالب.", "success" if updated else "warning")
    return redirect(request.referrer or url_for("finance_dashboard"))

@app.post("/finance/in-person/add")
@finance_login_required
def finance_add_in_person_student():
    full_name = request.form.get("full_name", "").strip()
    phone = request.form.get("phone", "").strip() or None
    course_name = request.form.get("course_name", "").strip()
    delivery_type = request.form.get("delivery_type", "near").strip()
    if delivery_type not in {"near", "remote"}:
        delivery_type = "near"
    monthly_amount = request.form.get("monthly_amount", "0").strip() or "0"
    start_month = request.form.get("start_month", current_month_label()).strip()
    notes = request.form.get("notes", "").strip() or None
    if not full_name or not course_name or not re.match(r"^\d{4}-\d{2}$", start_month):
        flash("اسم الطالب، الدورة، وشهر البداية مطلوبة.", "danger")
        return redirect(url_for("finance_dashboard"))
    try:
        amount = float(monthly_amount)
    except ValueError:
        amount = 0
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO in_person_students
                (full_name, phone, course_name, monthly_amount, start_month, delivery_type, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (full_name, phone, course_name, amount, start_month, delivery_type, notes))
        conn.commit()
        cur.close()
    if delivery_type == "remote":
        flash("تمت إضافة طالب عن بعد.", "success")
        return redirect(url_for("finance_dashboard", view="remote", category=course_name, month=start_month) + "#remote-students")
    flash("تمت إضافة طالب عن قرب.", "success")
    return redirect(url_for("finance_dashboard", view="near", near_category=course_name, month=start_month) + "#near-students")

@app.post("/finance/in-person/<int:student_id>/pay")
@finance_login_required
def finance_add_in_person_payment(student_id):
    month_label = request.form.get("month_label", current_month_label()).strip()
    amount = request.form.get("amount", "").strip()
    near_category = request.form.get("near_category", "").strip()
    return_view = request.form.get("return_view", "near").strip()
    if not re.match(r"^\d{4}-\d{2}$", month_label):
        flash("الشهر غير صحيح.", "danger")
        return redirect(url_for("finance_dashboard"))
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT monthly_amount FROM in_person_students WHERE id=%s AND active=TRUE", (student_id,))
        student = cur.fetchone()
        if not student:
            cur.close()
            flash("لم يتم العثور على الطالب.", "warning")
            if return_view == "remote":
                return redirect(url_for("finance_dashboard", view="remote", category=near_category, month=month_label) + "#remote-students")
            return redirect(url_for("finance_dashboard", view="near", near_category=near_category, month=month_label) + "#near-students")
        try:
            payment_amount = float(amount) if amount else float(student.get("monthly_amount") or 0)
        except ValueError:
            payment_amount = float(student.get("monthly_amount") or 0)
        cur.execute("""
            INSERT INTO in_person_payments (student_id, month_label, amount, recorded_by)
            VALUES (%s,%s,%s,%s)
            ON CONFLICT (student_id, month_label)
            DO UPDATE SET amount=EXCLUDED.amount,
                          paid_at=CURRENT_TIMESTAMP,
                          recorded_by=EXCLUDED.recorded_by
        """, (student_id, month_label, payment_amount, session.get("user_id")))
        conn.commit()
        cur.close()
    flash("تم تسجيل دفع الشهر.", "success")
    if return_view == "remote":
        return redirect(url_for("finance_dashboard", view="remote", category=near_category, month=month_label) + "#remote-students")
    return redirect(url_for("finance_dashboard", view="near", near_category=near_category, month=month_label) + "#near-students")

@app.post("/finance/in-person/<int:student_id>/deactivate")
@finance_login_required
def finance_deactivate_in_person_student(student_id):
    near_category = request.form.get("near_category", "").strip()
    return_view = request.form.get("return_view", "near").strip()
    with db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE in_person_students SET active=FALSE WHERE id=%s", (student_id,))
        conn.commit()
        cur.close()
    flash("تم إخفاء الطالب من قائمة عن قرب.", "warning")
    if return_view == "remote":
        flash("تم إخفاء الطالب من قائمة عن بعد.", "warning")
        return redirect(url_for("finance_dashboard", view="remote", category=near_category) + "#remote-students")
    return redirect(url_for("finance_dashboard", view="near", near_category=near_category) + "#near-students")

@app.post("/finance/categories/add")
@finance_login_required
def finance_add_category():
    name = request.form.get("name", "").strip()
    category_type = request.form.get("category_type", "").strip()
    if not name or category_type not in {"course", "section"}:
        flash("اسم القسم أو الدورة والنوعية مطلوبة.", "danger")
        return redirect(url_for("finance_dashboard") + "#finance-categories")
    try:
        with db() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO finance_categories (name, category_type, active)
                VALUES (%s,%s,TRUE)
                ON CONFLICT (name, category_type)
                DO UPDATE SET active=TRUE
            """, (name, category_type))
            conn.commit()
            cur.close()
        flash("تمت إضافة القسم أو الدورة.", "success")
    except IntegrityError:
        flash("هذا الاسم موجود بالفعل.", "warning")
    return redirect(url_for("finance_dashboard") + "#finance-categories")

@app.post("/finance/categories/<int:category_id>/delete")
@finance_login_required
def finance_delete_category(category_id):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE finance_categories SET active=FALSE WHERE id=%s", (category_id,))
        updated = cur.rowcount
        conn.commit()
        cur.close()
    flash("تم حذف الدورة أو القسم." if updated else "لم يتم العثور على العنصر.", "warning" if updated else "danger")
    return redirect(url_for("finance_dashboard") + "#finance-categories")

@app.route("/admin/lessons/create", methods=["POST"])
@admin_login_required
def admin_create_lesson():
    level = request.form.get("level","").strip()
    subject = request.form.get("subject","").strip()
    chapter = request.form.get("chapter_title","").strip()
    vfile = request.files.get("video")
    pfile = request.files.get("pdf")
    vurl = video_embed_url(request.form.get("video_url","")) or None

    if not level or not subject or not chapter:
        flash("La formation, la matiere et le titre sont requis.", "danger")
        return redirect(url_for("admin_dashboard"))

    allowed_subjects = {row["subject"] for row in fetch_course_subjects(level)}
    if subject not in allowed_subjects:
        flash("Matiere invalide pour cette formation.", "danger")
        return redirect(url_for("admin_dashboard"))

    vname = None
    pname = None
    if vfile and vfile.filename:
        if not allowed(vfile.filename, VIDEO_EXT):
            flash("Fichier video non autorise.", "danger")
            return redirect(url_for("admin_dashboard"))
        base = secure_filename(vfile.filename)
        vname = f"admin_{session['user_id']}_{int(time.time())}_{base}"
        vfile.save(os.path.join(VID_DIR, vname))
    if pfile and pfile.filename:
        if not allowed(pfile.filename, PDF_EXT):
            if vname:
                remove_upload(VID_DIR, vname)
            flash("Fichier PDF non autorise.", "danger")
            return redirect(url_for("admin_dashboard"))
        base = secure_filename(pfile.filename)
        pname = f"admin_{session['user_id']}_{int(time.time())}_{base}"
        pfile.save(os.path.join(PDF_DIR, pname))

    with db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO lessons
                       (subject, chapter_title, level, video_file, pdf_file, video_url, uploaded_by)
                       VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                    (subject, chapter, level, vname, pname, vurl, session["user_id"]))
        conn.commit(); cur.close()
    flash("Lecon ajoutee.", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/lessons/delete/<int:lesson_id>", methods=["POST"])
@admin_login_required
def admin_delete_lesson(lesson_id):
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT video_file, pdf_file FROM lessons WHERE id=%s", (lesson_id,))
        lesson = cur.fetchone()
        if not lesson:
            cur.close()
            flash("Lecon introuvable.", "danger")
            return redirect(url_for("admin_dashboard"))
        cur.execute("DELETE FROM lessons WHERE id=%s", (lesson_id,))
        conn.commit(); cur.close()

    remove_upload(VID_DIR, lesson["video_file"])
    remove_upload(PDF_DIR, lesson["pdf_file"])
    flash("Lecon supprimee.", "warning")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/free-pdfs/create", methods=["POST"])
@admin_login_required
def admin_create_free_pdf():
    course_code = request.form.get("course_code","").strip()
    subject = request.form.get("subject","").strip()
    title = request.form.get("title","").strip()
    drive_url = request.form.get("drive_url","").strip()
    sort_order = request.form.get("sort_order","0").strip()

    if not course_code or not subject or not title or not drive_url:
        flash("La formation, la matière, le titre et le lien PDF sont requis.", "danger")
        return redirect(url_for("admin_dashboard"))
    if "drive.google.com" not in drive_url:
        flash("Veuillez utiliser un lien Google Drive.", "danger")
        return redirect(url_for("admin_dashboard"))
    try:
        sort_order = int(sort_order)
    except ValueError:
        sort_order = 0

    allowed_subjects = {row["subject"] for row in fetch_course_subjects(course_code)}
    if subject not in allowed_subjects:
        flash("Matière invalide pour cette formation.", "danger")
        return redirect(url_for("admin_dashboard"))

    with db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO free_pdfs (course_code, subject, title, drive_url, sort_order, active)
                       VALUES (%s,%s,%s,%s,%s,TRUE)""",
                    (course_code, subject, title, drive_url, sort_order))
        conn.commit(); cur.close()
    flash("PDF gratuit ajouté.", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/free-pdfs/delete/<int:pdf_id>")
@admin_login_required
def admin_delete_free_pdf(pdf_id):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM free_pdfs WHERE id=%s", (pdf_id,))
        conn.commit(); cur.close()
    flash("PDF gratuit supprimé.", "warning")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/courses/create", methods=["POST"])
@admin_login_required
def admin_create_course():
    code = request.form.get("code","").strip().upper()
    title = request.form.get("title","").strip()
    subtitle = request.form.get("subtitle","").strip()
    description = request.form.get("description","").strip()
    badge = request.form.get("badge","").strip()
    icon = request.form.get("icon","").strip() or "ðŸ“˜"
    theme = request.form.get("theme","blue").strip() or "blue"
    sort_order = request.form.get("sort_order","0").strip()

    if not code or not title:
        flash("Le code et le titre de la formation sont requis.", "danger")
        return redirect(url_for("admin_dashboard"))
    try:
        sort_order = int(sort_order)
    except ValueError:
        sort_order = 0

    try:
        ensure_courses_table()
        with db() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO courses (code, title, subtitle, description, badge, icon, theme, sort_order, active)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
            """, (code, title, subtitle, description, badge, icon, theme, sort_order))
            conn.commit(); cur.close()
        flash("Formation ajoutÃ©e.", "success")
    except IntegrityError:
        flash("Ce code de formation existe dÃ©jÃ .", "danger")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/courses/subjects/create", methods=["POST"])
@admin_login_required
def admin_create_course_subject():
    course_code = request.form.get("course_code","").strip()
    subject = request.form.get("subject","").strip()
    sort_order = request.form.get("sort_order","0").strip()
    if not course_code or not subject:
        flash("La formation et la matiÃ¨re sont requises.", "danger")
        return redirect(url_for("admin_dashboard"))
    try:
        sort_order = int(sort_order)
    except ValueError:
        sort_order = 0
    ensure_courses_table()
    try:
        with db() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO course_subjects (course_code, subject, sort_order)
                VALUES (%s,%s,%s)
            """, (course_code, subject, sort_order))
            conn.commit(); cur.close()
        flash("MatiÃ¨re ajoutÃ©e.", "success")
    except IntegrityError:
        flash("Cette matiÃ¨re existe dÃ©jÃ  pour cette formation.", "warning")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/courses/subjects/delete/<int:sid>")
@admin_login_required
def admin_delete_course_subject(sid):
    ensure_courses_table()
    with db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM course_subjects WHERE id=%s", (sid,))
        conn.commit(); cur.close()
    flash("MatiÃ¨re supprimÃ©e.", "warning")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/users/<int:uid>/assign-teacher", methods=["POST"])
@admin_login_required
def admin_assign_teacher(uid):
    level = request.form.get("level","").strip()
    subject = request.form.get("subject","").strip()
    if not level or not subject:
        flash("La formation et la matiÃ¨re sont requises.", "danger")
        return redirect(url_for("admin_dashboard"))
    allowed_subjects = {row["subject"] for row in fetch_course_subjects(level)}
    if subject not in allowed_subjects:
        return redirect(url_for("admin_dashboard"))
    with db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE users SET level=%s, subject=%s WHERE id=%s AND role='teacher'",
                    (level, subject, uid))
        conn.commit(); cur.close()
    flash("Compte enseignant mis Ã  jour.", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/courses/delete/<int:cid>")
@admin_login_required
def admin_delete_course(cid):
    ensure_courses_table()
    with db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM courses WHERE id=%s", (cid,))
        conn.commit(); cur.close()
    flash("Formation supprimÃ©e.", "warning")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/create-teacher", methods=["POST"])
@admin_login_required
def admin_create_teacher():
    t_user  = request.form.get("t_username","").strip()
    t_phone = request.form.get("t_phone","").strip()
    t_pass  = request.form.get("t_password","")
    t_level = request.form.get("t_level") or None
    t_subject = request.form.get("t_subject","").strip() or None
    if not t_user or not t_phone or not t_pass or not t_level or not t_subject:
        flash("Nom dâ€™utilisateur, numÃ©ro et mot de passe sont requis.", "danger")
        return redirect(url_for("admin_dashboard"))
    try:
        with db() as conn:
            cur = conn.cursor()
            cur.execute("""INSERT INTO users (username, phone, password, role, level, subject, status, phone_verified)
                           VALUES (%s,%s,%s,'teacher',%s,%s,'active',TRUE)""",
                        (t_user, t_phone, hash_password(t_pass), t_level, t_subject))
            conn.commit(); cur.close()
        flash("Compte enseignant crÃ©Ã©.", "success")
    except IntegrityError as e:
        msg = "Nom dâ€™utilisateur dÃ©jÃ  utilisÃ©."
        if "phone" in str(e): msg = "NumÃ©ro de tÃ©lÃ©phone dÃ©jÃ  utilisÃ©."
        flash(msg, "danger")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/create-admin", methods=["POST"])
@developer_required
def developer_create_admin():
    username = request.form.get("admin_username","").strip()
    phone = request.form.get("admin_phone","").strip()
    password = request.form.get("admin_password","")
    if not username or not phone or not password:
        flash("Nom d'utilisateur, numéro et mot de passe sont requis.", "danger")
        return redirect(url_for("admin_dashboard"))
    try:
        with db() as conn:
            cur = conn.cursor()
            cur.execute("""INSERT INTO users (username, phone, password, role, status, phone_verified)
                           VALUES (%s,%s,%s,'admin','active',TRUE)""",
                        (username, phone, hash_password(password)))
            conn.commit(); cur.close()
        flash("Compte admin créé.", "success")
    except IntegrityError as e:
        msg = "Nom d'utilisateur déjà utilisé."
        if "phone" in str(e): msg = "Numéro de téléphone déjà utilisé."
        flash(msg, "danger")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/create-finance", methods=["POST"])
@developer_required
def developer_create_finance():
    username = request.form.get("finance_username","").strip()
    phone = request.form.get("finance_phone","").strip()
    password = request.form.get("finance_password","")
    if not username or not phone or not password:
        flash("اسم حساب المالية، الرقم وكلمة المرور مطلوبة.", "danger")
        return redirect(url_for("admin_dashboard"))
    try:
        with db() as conn:
            cur = conn.cursor()
            cur.execute("""INSERT INTO users
                           (username, phone, password, role, status, phone_verified, payment_status)
                           VALUES (%s,%s,%s,'finance','active',TRUE,'not_applicable')""",
                        (username, phone, hash_password(password)))
            conn.commit(); cur.close()
        flash("تم إنشاء حساب المالية.", "success")
    except IntegrityError as e:
        msg = "اسم المستخدم مستخدم بالفعل."
        if "phone" in str(e): msg = "رقم الهاتف مستخدم بالفعل."
        flash(msg, "danger")
    return redirect(url_for("admin_dashboard"))

# ===== Tableau de bord Enseignant =====
@app.route("/teacher", methods=["GET","POST"])
@login_required("teacher")
def teacher_dashboard():
    courses = fetch_courses(active_only=True)
    assigned_level = session.get("level")
    assigned_subject = session.get("subject")
    if request.method == "POST":
        subject = assigned_subject
        chapter = request.form.get("chapter_title","").strip()
        level   = assigned_level
        vfile   = request.files.get("video")
        pfile   = request.files.get("pdf")
        vurl    = video_embed_url(request.form.get("video_url","")) or None

        if not subject or not chapter or not level:
            flash("Champs obligatoires manquants.", "danger"); return redirect(url_for("teacher_dashboard"))

        vname = None; pname = None
        if vfile and vfile.filename and allowed(vfile.filename, VIDEO_EXT):
            base = secure_filename(vfile.filename)
            vname = f"{session['user_id']}_{int(time.time())}_{base}"
            vfile.save(os.path.join(VID_DIR, vname))
        if pfile and pfile.filename and allowed(pfile.filename, PDF_EXT):
            base = secure_filename(pfile.filename)
            pname = f"{session['user_id']}_{int(time.time())}_{base}"
            pfile.save(os.path.join(PDF_DIR, pname))

        with db() as conn:
            cur = conn.cursor()
            # NEW: insert video_url
            cur.execute("""INSERT INTO lessons (subject,chapter_title,level,video_file,pdf_file,video_url,uploaded_by)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (subject, chapter, level, vname, pname, vurl, session["user_id"]))
            conn.commit(); cur.close()
        flash("LeÃ§on ajoutÃ©e.", "success")
        return redirect(url_for("teacher_dashboard"))

    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT * FROM lessons
                       WHERE uploaded_by=%s AND level=%s AND subject=%s
                       ORDER BY uploaded_at DESC""",
                    (session["user_id"], assigned_level, assigned_subject))
        my_lessons = cur.fetchall(); cur.close()
    return render_template("teacher.html", lessons=my_lessons, courses=courses,
                           assigned_level=assigned_level, assigned_subject=assigned_subject)

@app.route("/teacher/lessons/<int:lesson_id>/edit", methods=["POST"])
@login_required("teacher")
def teacher_edit_lesson(lesson_id):
    assigned_level = session.get("level")
    assigned_subject = session.get("subject")
    chapter = request.form.get("chapter_title","").strip()
    vurl = video_embed_url(request.form.get("video_url","")) or None
    pfile = request.files.get("pdf")
    remove_pdf = request.form.get("remove_pdf") == "1"

    if not chapter:
        flash("Titre de la leçon requis.", "danger")
        return redirect(url_for("teacher_dashboard"))

    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT id, pdf_file FROM lessons
                       WHERE id=%s AND uploaded_by=%s AND level=%s AND subject=%s""",
                    (lesson_id, session["user_id"], assigned_level, assigned_subject))
        lesson = cur.fetchone()
        if not lesson:
            cur.close()
            flash("Leçon introuvable ou non autorisée.", "danger")
            return redirect(url_for("teacher_dashboard"))

        if pfile and pfile.filename:
            if not allowed(pfile.filename, PDF_EXT):
                cur.close()
                flash("Fichier PDF non autorisé.", "danger")
                return redirect(url_for("teacher_dashboard"))
        new_pdf = lesson["pdf_file"]
        if remove_pdf:
            remove_upload(PDF_DIR, lesson["pdf_file"])
            new_pdf = None
        if pfile and pfile.filename:
            remove_upload(PDF_DIR, lesson["pdf_file"])
            base = secure_filename(pfile.filename)
            new_pdf = f"{session['user_id']}_{int(time.time())}_{base}"
            pfile.save(os.path.join(PDF_DIR, new_pdf))

        cur.execute("""UPDATE lessons
                       SET chapter_title=%s, video_url=%s, pdf_file=%s
                       WHERE id=%s AND uploaded_by=%s""",
                    (chapter, vurl, new_pdf, lesson_id, session["user_id"]))
        conn.commit()
        cur.close()
    flash("Leçon mise à jour.", "success")
    return redirect(url_for("teacher_dashboard"))

@app.route("/teacher/lessons/<int:lesson_id>/delete", methods=["POST"])
@login_required("teacher")
def teacher_delete_lesson(lesson_id):
    assigned_level = session.get("level")
    assigned_subject = session.get("subject")
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT id, video_file, pdf_file FROM lessons
                       WHERE id=%s AND uploaded_by=%s AND level=%s AND subject=%s""",
                    (lesson_id, session["user_id"], assigned_level, assigned_subject))
        lesson = cur.fetchone()
        if not lesson:
            cur.close()
            flash("Leçon introuvable ou non autorisée.", "danger")
            return redirect(url_for("teacher_dashboard"))
        cur.execute("DELETE FROM lessons WHERE id=%s AND uploaded_by=%s",
                    (lesson_id, session["user_id"]))
        conn.commit()
        cur.close()

    remove_upload(VID_DIR, lesson["video_file"])
    remove_upload(PDF_DIR, lesson["pdf_file"])
    flash("Leçon supprimée.", "warning")
    return redirect(url_for("teacher_dashboard"))

# ===== Tableau de bord Ã‰tudiant =====
@app.route("/student")
@login_required("student")
def student_dashboard():
    level = session.get("level")
    subject = request.args.get("subject")  # NEW: optional filter
    subjects = fetch_course_subjects(level) if level else []
    allowed_subjects = {row["subject"] for row in subjects}
    if subject and subject not in allowed_subjects:
        subject = None
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT phone, selected_subjects FROM users WHERE id=%s", (session["user_id"],))
        student = cur.fetchone()
        selected_subjects = parse_selected_subjects(student.get("selected_subjects") if student else None)
        if subject:
            cur.execute("""SELECT * FROM lessons
                           WHERE level=%s AND subject=%s
                           ORDER BY uploaded_at DESC""",
                        (level, subject))
        elif selected_subjects:
            cur.execute("""SELECT * FROM lessons
                           WHERE level=%s AND subject = ANY(%s)
                           ORDER BY uploaded_at DESC""",
                        (level, selected_subjects))
        else:
            cur.execute("""SELECT * FROM lessons
                           WHERE level=%s
                           ORDER BY CASE subject
                                      WHEN 'Math' THEN 1
                                      WHEN 'Physique' THEN 2
                                      WHEN 'Chimie' THEN 3
                                      WHEN 'Science naturelle' THEN 4
                                      ELSE 5
                                    END,
                                    uploaded_at DESC""",
                        (level,))
        lessons = cur.fetchall(); cur.close()
    return render_template("student.html", lessons=lessons, level=level,
                           subjects=subjects, student=student)

# ===== Mobile/API Backend =====
api_serializer = URLSafeTimedSerializer(app.secret_key, salt="epsilon-mobile-api")
API_TOKEN_MAX_AGE = 60 * 60 * 24 * 30

def api_error(message, status=400, code=None):
    payload = {"error": code or message, "message": message}
    return jsonify(payload), status

def api_user_payload(user):
    return {
        "id": str(user["id"]),
        "name": user["username"],
        "username": user["username"],
        "email": user["phone"],
        "phone": user["phone"],
        "role": user["role"],
        "status": user["status"],
        "paymentStatus": user.get("payment_status"),
        "classId": user.get("level"),
        "courseId": user.get("level"),
        "level": user.get("level"),
        "subject": user.get("subject"),
        "selectedSubjects": parse_selected_subjects(user.get("selected_subjects")),
        "paymentSenderPhone": user.get("payment_sender_phone"),
        "paymentProofUrl": url_for("static", filename=f"uploads/payments/{user['payment_image']}", _external=True)
        if user.get("payment_image") else None,
        "createdAt": user.get("created_at").isoformat() if user.get("created_at") else None,
    }

def api_course_payload(course, subjects_by_course=None):
    subjects_by_course = subjects_by_course or {}
    subjects = subjects_by_course.get(course["code"])
    if subjects is None:
        subjects = fetch_course_subjects(course["code"])
    return {
        "id": course["code"],
        "dbId": str(course["id"]),
        "code": course["code"],
        "title": course["title"],
        "name": course["title"],
        "level": course["code"],
        "classId": course["code"],
        "description": course.get("description") or course.get("subtitle") or "",
        "price": course.get("badge") or "",
        "subjects": [row["subject"] for row in subjects],
        "isActive": bool(course.get("active")),
        "sortOrder": course.get("sort_order") or 0,
    }

def api_lesson_payload(lesson):
    video_url = lesson.get("video_url")
    if not video_url and lesson.get("video_file"):
        video_url = url_for("static", filename=f"uploads/videos/{lesson['video_file']}", _external=True)
    pdf_url = url_for("static", filename=f"uploads/pdfs/{lesson['pdf_file']}", _external=True) if lesson.get("pdf_file") else None
    return {
        "id": str(lesson["id"]),
        "title": lesson["chapter_title"],
        "url": video_url or pdf_url or "",
        "videoUrl": video_url,
        "pdfUrl": pdf_url,
        "teacherId": str(lesson["uploaded_by"]) if lesson.get("uploaded_by") else None,
        "classId": lesson["level"],
        "courseId": lesson["level"],
        "level": lesson["level"],
        "subject": lesson["subject"],
        "isPublished": True,
        "createdAt": lesson["uploaded_at"].isoformat() if lesson.get("uploaded_at") else None,
    }

RESULT_EXAM_TYPES = {"concours", "brevet", "bac-first"}
RESULT_EXAM_LABELS = {
    "concours": "كونكور",
    "brevet": "ابريفة",
    "bac-first": "الباكالوريا الدورة الأولى",
}
RESULT_FIELD_ALIASES = {
    "candidate_number": [
        "رقم الجلوس", "رقم", "numero", "num", "matricule", "nni", "candidate number",
    ],
    "full_name": ["الاسم", "الإسم", "اسم", "nom", "name", "full name", "candidat"],
    "birth_place": ["محل الميلاد", "مكان الميلاد", "lieu naissance", "place of birth"],
    "birth_date": ["تاريخ الميلاد", "date naissance", "date de naissance", "birth date"],
    "wilaya": ["الولاية", "wilaya"],
    "moughataa": ["المقاطعة", "moughataa", "departement", "département"],
    "center_name": ["centre examen", "centre examen fr", "center", "centre", "مركز", "مركز الامتحان"],
    "score": ["moy bac", "moy", "moyenne", "المعدل", "النتيجة", "score"],
    "decision": ["قرار", "القرار", "decision", "décision", "resultat", "résultat"],
    "rank": ["الرتبة", "الترتيب", "rang", "rank", "classement"],
}

def normalize_excel_text(value):
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[_\-/]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def canonical_result_field(header):
    normalized = normalize_excel_text(header)
    for field, aliases in RESULT_FIELD_ALIASES.items():
        for alias in aliases:
            alias_norm = normalize_excel_text(alias)
            if normalized == alias_norm or alias_norm in normalized:
                return field
    return None

def excel_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text

def result_score_text(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            number = int(value)
            if number == 0:
                return "0"
            if abs(number) >= 100000:
                return f"{number / 1000000:.6f}".rstrip("0").rstrip(".").replace(".", ",")
        return str(value).replace(".", ",")
    return excel_cell_text(value)

def normalize_candidate_number(value):
    digits = re.sub(r"\D+", "", str(value or ""))
    normalized = digits.lstrip("0")
    return normalized or ("0" if digits else "")

def concours_number_match_sql():
    return """
        (
            COALESCE(NULLIF(regexp_replace(candidate_number, '\\D', '', 'g'), ''), '0')::bigint = %s::bigint
            OR COALESCE(NULLIF(regexp_replace(COALESCE(raw_data->>'Numéro Ins', raw_data->>'Numero Ins', ''), '\\D', '', 'g'), ''), '0')::bigint = %s::bigint
        )
    """

def normalize_concours_name_query(value):
    text = normalize_excel_text(value)
    return re.sub(r"[\s/ـ]+", "", text)

def search_concours_by_name(cur, query, center_name=None, limit=20):
    normalized_query = normalize_concours_name_query(query)
    center_clause = "AND center_name = %s" if center_name else ""
    center_params = [center_name] if center_name else []
    cur.execute("""
        SELECT *
        FROM national_exam_results
        WHERE exam_type='concours'
          {center_clause}
          AND (
              full_name ILIKE %s
              OR regexp_replace(lower(full_name), '[[:space:]/ـ]+', '', 'g') ILIKE %s
          )
        ORDER BY
          CASE
            WHEN full_name = %s THEN 0
            WHEN lower(full_name) = lower(%s) THEN 1
            WHEN regexp_replace(lower(full_name), '[[:space:]/ـ]+', '', 'g') = %s THEN 2
            WHEN full_name ILIKE %s THEN 3
            ELSE 4
          END,
          length(full_name) ASC,
          full_name ASC
        LIMIT %s
    """.format(center_clause=center_clause), (
        *center_params,
        f"%{query}%",
        f"%{normalized_query}%",
        query,
        query,
        normalized_query,
        f"{query}%",
        limit,
    ))

def db_text(value, max_length=None):
    text = excel_cell_text(value)
    if not text:
        return None
    if max_length and len(text) > max_length:
        return text[:max_length]
    return text

def parse_results_workbook(file_storage):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed.") from exc

    workbook = load_workbook(file_storage, read_only=True, data_only=True)
    sheet = workbook.active
    header_row_number = None
    field_by_index = {}
    header_by_index = {}

    sheet_max_row = sheet.max_row if isinstance(sheet.max_row, int) and sheet.max_row > 0 else 50
    scan_until_row = min(sheet_max_row, 50)
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=scan_until_row, values_only=True),
        start=1,
    ):
        mapped = {}
        labels = {}
        for index, value in enumerate(row):
            label = excel_cell_text(value)
            field = canonical_result_field(label)
            if field and field not in mapped.values():
                mapped[index] = field
                labels[index] = label
        if len(mapped) >= 2 and "full_name" in mapped.values():
            header_row_number = row_number
            field_by_index = mapped
            header_by_index = labels
            break

    if not header_row_number:
        raise ValueError("لم يتم العثور على صف عناوين واضح داخل ملف Excel.")

    results = []
    start_row = header_row_number + 1

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        parsed = {}
        raw = {}
        for index, value in enumerate(row):
            text = excel_cell_text(value)
            if index in header_by_index:
                raw[header_by_index[index]] = text
            field = field_by_index.get(index)
            if field and text:
                if field == "score":
                    text = result_score_text(value)
                parsed[field] = text
        if not parsed.get("full_name") and not parsed.get("candidate_number"):
            continue
        if not parsed.get("full_name"):
            continue
        parsed["raw_data"] = raw
        results.append(parsed)
    workbook.close()
    return results

def insert_national_results(exam_type, filename, parsed_rows, uploaded_by):
    filename = db_text(filename, 255) or "results.xlsx"
    values = [
        (
            exam_type,
            db_text(row.get("candidate_number"), 80),
            db_text(row.get("full_name"), 255),
            db_text(row.get("birth_place"), 160),
            db_text(row.get("birth_date"), 80),
            db_text(row.get("wilaya"), 160),
            db_text(row.get("moughataa"), 160),
            db_text(row.get("center_name"), 255),
            db_text(row.get("score"), 80),
            db_text(row.get("decision"), 160),
            db_text(row.get("rank"), 80),
            json.dumps(row.get("raw_data") or {}, ensure_ascii=False),
        )
        for row in parsed_rows
        if db_text(row.get("full_name"), 255)
    ]
    if not values:
        raise ValueError("لم يتم العثور على أسماء صالحة داخل ملف Excel.")
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO result_uploads (exam_type, original_filename, rows_imported, uploaded_by)
            VALUES (%s,%s,%s,%s)
            RETURNING id, uploaded_at
        """, (exam_type, filename, len(values), uploaded_by))
        upload_id, uploaded_at = cur.fetchone()
        cur.execute("DELETE FROM national_exam_results WHERE exam_type=%s", (exam_type,))
        buffer = StringIO()
        writer = csv.writer(buffer)
        for row in values:
            writer.writerow(row + (upload_id,))
        buffer.seek(0)
        cur.copy_expert("""
            COPY national_exam_results
                (exam_type, candidate_number, full_name, birth_place, birth_date,
                 wilaya, moughataa, center_name, score, decision, rank, raw_data, upload_id)
            FROM STDIN WITH (FORMAT CSV)
        """, buffer)
        conn.commit()
        cur.close()
    return {"id": upload_id, "uploaded_at": uploaded_at}

def api_result_payload(row):
    raw_data = row.get("raw_data") or {}
    candidate_number = row.get("candidate_number") or ""
    if row.get("exam_type") == "concours":
        candidate_number = (
            raw_data.get("Numéro Ins")
            or raw_data.get("Numero Ins")
            or candidate_number
        )
    return {
        "id": str(row["id"]),
        "examType": row["exam_type"],
        "candidateNumber": candidate_number,
        "fullName": row.get("full_name") or "",
        "birthPlace": row.get("birth_place") or "",
        "birthDate": row.get("birth_date") or "",
        "wilaya": row.get("wilaya") or "",
        "moughataa": row.get("moughataa") or "",
        "centerName": row.get("center_name") or "",
        "score": row.get("score") or "",
        "decision": row.get("decision") or "",
        "rank": row.get("rank") or "",
        "rawData": raw_data,
    }

def api_current_user():
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return None
    token = auth_header.removeprefix("Bearer ").strip()
    try:
        data = api_serializer.loads(token, max_age=API_TOKEN_MAX_AGE)
    except (BadSignature, SignatureExpired):
        return None
    user_id = data.get("user_id")
    if not user_id:
        return None
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("SELECT * FROM users WHERE id=%s", (user_id,))
        user = cur.fetchone()
        cur.close()
    return user

def api_login_required(role=None):
    def deco(fn):
        @wraps(fn)
        def wrap(*args, **kwargs):
            user = api_current_user()
            if not user:
                return api_error("Authentication is required.", 401, "unauthenticated")
            if user["status"] != "active" and user["role"] not in ADMIN_ROLES:
                return api_error("Account is not active.", 403, "account_inactive")
            if role:
                allowed = role if isinstance(role, (set, tuple, list)) else {role}
                if user["role"] not in allowed:
                    return api_error("Permission denied.", 403, "permission_denied")
            request.api_user = user
            return fn(*args, **kwargs)
        return wrap
    return deco

@app.get("/api/health")
def api_health():
    return jsonify({"ok": True, "database": bool(DATABASE_URL), "service": "epsilon-flask"})

@app.post("/api/auth/login")
def api_login():
    data = request.get_json(silent=True) or {}
    identifier = (data.get("identifier") or data.get("username") or data.get("phone") or data.get("email") or "").strip()
    password = data.get("password") or ""
    if not identifier or not password:
        return api_error("Phone number and password are required.", 400, "missing_credentials")

    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT * FROM users
                       WHERE (username=%s OR phone=%s) AND password=%s
                       LIMIT 1""", (identifier, identifier, hash_password(password)))
        user = cur.fetchone()
        cur.close()

    if not user:
        return api_error("Invalid phone number or password.", 401, "invalid_credentials")
    if user["status"] not in {"active", "pending"}:
        return api_error("Account is blocked.", 403, "account_blocked")

    token = api_serializer.dumps({"user_id": user["id"], "role": user["role"]})
    return jsonify({"token": token, "user": api_user_payload(user)})

@app.get("/api/me")
@api_login_required()
def api_me():
    return jsonify({"user": api_user_payload(request.api_user)})

@app.get("/api/users")
@api_login_required(ADMIN_ROLES)
def api_users():
    with db() as conn:
        cur = dict_cursor(conn)
        if request.api_user["role"] == "developer":
            cur.execute("SELECT * FROM users ORDER BY id DESC")
        else:
            cur.execute("SELECT * FROM users WHERE role <> 'developer' ORDER BY id DESC")
        users = cur.fetchall()
        cur.close()
    return jsonify({"users": [api_user_payload(user) for user in users]})

@app.patch("/api/users/<int:user_id>/status")
@api_login_required(ADMIN_ROLES)
def api_update_user_status(user_id):
    data = request.get_json(silent=True) or {}
    status = (data.get("status") or "").strip()
    if status not in {"pending", "active", "blocked", "rejected"}:
        return api_error("Unsupported account status.", 400, "invalid_status")
    if request.api_user["role"] != "developer" and target_is_developer(user_id):
        return api_error("Permission denied.", 403, "permission_denied")
    with db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE users SET status=%s WHERE id=%s", (status, user_id))
        conn.commit()
        cur.close()
    return jsonify({"id": str(user_id), "status": status})

@app.delete("/api/users/<int:user_id>")
@api_login_required(ADMIN_ROLES)
def api_delete_user(user_id):
    if user_id == request.api_user["id"]:
        return api_error("You cannot delete your own account.", 400, "cannot_delete_self")
    if request.api_user["role"] != "developer" and target_is_developer(user_id):
        return api_error("Permission denied.", 403, "permission_denied")
    result = delete_user_with_teacher_lessons(user_id)
    return jsonify({
        "deleted": result["user_deleted"],
        "id": str(user_id),
        "lessonsDeleted": result["lessons_deleted"],
    })

@app.post("/api/admin/users")
@api_login_required(ADMIN_ROLES)
def api_create_user():
    data = request.get_json(silent=True) or {}
    username = (data.get("name") or data.get("username") or "").strip()
    phone = (data.get("phone") or data.get("email") or "").strip()
    password = data.get("password") or ""
    role = (data.get("role") or "student").strip()
    level = data.get("level") or data.get("classId") or data.get("courseId")
    subject = data.get("subject")
    payment_sender_phone = (data.get("paymentSenderPhone") or "").strip() or None
    status = data.get("status") or ("active" if role in {"teacher", "admin"} else "pending")
    if role not in {"admin", "teacher", "student"}:
        return api_error("Unsupported role.", 400, "invalid_role")
    if role == "admin" and request.api_user["role"] != "developer":
        return api_error("Only developer can create admins.", 403, "permission_denied")
    if not username or not phone or not password:
        return api_error("Name, phone number and password are required.", 400, "missing_fields")

    try:
        with db() as conn:
            cur = dict_cursor(conn)
            cur.execute("""INSERT INTO users
                           (username, phone, password, role, level, subject, payment_sender_phone, status, phone_verified)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
                           RETURNING *""",
                        (username, phone, hash_password(password), role, level, subject, payment_sender_phone, status))
            user = cur.fetchone()
            conn.commit()
            cur.close()
    except IntegrityError:
        return api_error("Name or phone number already exists.", 409, "user_exists")

    return jsonify({"user": api_user_payload(user)}), 201

@app.post("/api/auth/register-student")
def api_register_student():
    data = request.get_json(silent=True) or {}
    username = (data.get("name") or data.get("username") or "").strip()
    phone = (data.get("phone") or data.get("email") or "").strip()
    password = data.get("password") or ""
    level = data.get("level") or data.get("classId") or data.get("courseId")
    subjects = parse_selected_subjects(data.get("selectedSubjects") or data.get("subjects"))
    payment_sender_phone = (data.get("paymentSenderPhone") or "").strip() or None
    if not username or not phone or not password or not level:
        return api_error("Name, phone number, password and course are required.", 400, "missing_fields")
    allowed_subjects = [row["subject"] for row in fetch_course_subjects(level)]
    selected_subjects = selected_subjects_json(subjects or allowed_subjects, allowed_subjects)
    try:
        with db() as conn:
            cur = dict_cursor(conn)
            cur.execute("""INSERT INTO users
                           (username, phone, password, role, level, selected_subjects, payment_sender_phone, status, phone_verified)
                           VALUES (%s,%s,%s,'student',%s,%s,%s,'pending',TRUE)
                           RETURNING *""",
                        (username, phone, hash_password(password), level, selected_subjects, payment_sender_phone))
            user = cur.fetchone()
            conn.commit()
            cur.close()
    except IntegrityError:
        return api_error("Name or phone number already exists.", 409, "user_exists")
    return jsonify({"user": api_user_payload(user)}), 201

@app.post("/api/admin/teachers")
@api_login_required(ADMIN_ROLES)
def api_create_teacher():
    data = request.get_json(silent=True) or {}
    data["role"] = "teacher"
    data["status"] = "active"
    return api_create_user()

@app.post("/api/admin/students")
@api_login_required(ADMIN_ROLES)
def api_create_student_by_admin():
    data = request.get_json(silent=True) or {}
    data["role"] = "student"
    data["status"] = "active"
    return api_create_user()

@app.get("/api/courses")
def api_courses():
    courses = fetch_courses(active_only=request.args.get("all") != "1")
    subjects_by_course = {}
    for subject in fetch_course_subjects():
        subjects_by_course.setdefault(subject["course_code"], []).append(subject)
    return jsonify({
        "courses": [
            api_course_payload(course, subjects_by_course)
            for course in courses
        ]
    })

@app.post("/api/courses")
@api_login_required(ADMIN_ROLES)
def api_create_course():
    data = request.get_json(silent=True) or {}
    title = (data.get("title") or "").strip()
    description = (data.get("description") or "").strip()
    price = (data.get("price") or "").strip()
    subjects = data.get("subjects") if isinstance(data.get("subjects"), list) else []
    if not title:
        return api_error("Title is required.", 400, "missing_title")
    code_base = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-") or f"course-{int(time.time())}"
    code = code_base
    with db() as conn:
        cur = dict_cursor(conn)
        suffix = 1
        while True:
            cur.execute("SELECT id FROM courses WHERE code=%s", (code,))
            if not cur.fetchone():
                break
            suffix += 1
            code = f"{code_base}-{suffix}"
        cur.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order FROM courses")
        sort_order = cur.fetchone()["next_order"]
        cur.execute("""INSERT INTO courses
                       (code, title, subtitle, description, badge, sort_order, active)
                       VALUES (%s,%s,%s,%s,%s,%s,TRUE)
                       RETURNING *""", (code, title, description, description, price, sort_order))
        course = cur.fetchone()
        for index, subject in enumerate([str(item).strip() for item in subjects if str(item).strip()], start=1):
            cur.execute("""INSERT INTO course_subjects (course_code, subject, sort_order)
                           VALUES (%s,%s,%s)
                           ON CONFLICT (course_code, subject) DO NOTHING""", (code, subject, index))
        conn.commit()
        cur.close()
    return jsonify({"course": api_course_payload(course)}), 201

@app.delete("/api/courses/<course_key>")
@api_login_required(ADMIN_ROLES)
def api_delete_course(course_key):
    with db() as conn:
        cur = conn.cursor()
        if course_key.isdigit():
            cur.execute("DELETE FROM courses WHERE id=%s", (int(course_key),))
        else:
            cur.execute("DELETE FROM courses WHERE code=%s", (course_key,))
        conn.commit()
        cur.close()
    return jsonify({"deleted": True, "id": str(course_key)})

@app.get("/api/classes")
def api_classes():
    courses = fetch_courses(active_only=True)
    classes = [{"id": course["code"], "name": course["title"], "level": course["code"]} for course in courses]
    return jsonify({"classes": classes})

@app.get("/api/settings")
def api_settings():
    return jsonify({"settings": fetch_app_settings()})

@app.patch("/api/settings")
@api_login_required(ADMIN_ROLES)
def api_update_settings():
    data = request.get_json(silent=True) or {}
    settings = save_app_settings({
        "paymentNumber": data.get("paymentNumber"),
        "paymentAmount": data.get("paymentAmount"),
    })
    return jsonify({"settings": settings})

@app.post("/api/classes")
@api_login_required(ADMIN_ROLES)
def api_create_class():
    data = request.get_json(silent=True) or {}
    title = (data.get("name") or data.get("level") or "").strip()
    if not title:
        return api_error("Name is required.", 400, "missing_name")
    data["title"] = title
    data["description"] = data.get("description") or data.get("level") or ""
    return api_create_course()

@app.get("/api/courses/<course_code>/subjects")
def api_course_subjects(course_code):
    return jsonify({"subjects": fetch_course_subjects(course_code)})

@app.get("/api/lessons")
@api_login_required()
def api_lessons():
    user = request.api_user
    level = request.args.get("level") or request.args.get("classId") or request.args.get("courseId")
    subject = request.args.get("subject")
    params = []
    clauses = []
    if user["role"] == "student":
        clauses.append("level=%s")
        params.append(user["level"])
        selected_subjects = parse_selected_subjects(user.get("selected_subjects"))
        if selected_subjects:
            clauses.append("subject = ANY(%s)")
            params.append(selected_subjects)
    elif user["role"] == "teacher":
        clauses.extend(["uploaded_by=%s", "level=%s", "subject=%s"])
        params.extend([user["id"], user["level"], user["subject"]])
    elif level:
        clauses.append("level=%s")
        params.append(level)
    if subject:
        clauses.append("subject=%s")
        params.append(subject)
    where = "WHERE " + " AND ".join(clauses) if clauses else ""
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute(f"SELECT * FROM lessons {where} ORDER BY uploaded_at DESC", params)
        lessons = cur.fetchall()
        cur.close()
    return jsonify({"lessons": [api_lesson_payload(lesson) for lesson in lessons]})

@app.post("/api/lessons")
@api_login_required({"admin", "developer", "teacher"})
def api_create_lesson():
    data = request.get_json(silent=True) or {}
    user = request.api_user
    title = (data.get("title") or data.get("chapter_title") or "").strip()
    level = data.get("level") or data.get("classId") or data.get("courseId") or user.get("level")
    subject = data.get("subject") or user.get("subject")
    video_url = video_embed_url(data.get("url") or data.get("videoUrl") or "")
    if not title or not level or not subject:
        return api_error("Title, level and subject are required.", 400, "missing_fields")
    if user["role"] == "teacher" and (level != user.get("level") or subject != user.get("subject")):
        return api_error("Permission denied.", 403, "permission_denied")
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""INSERT INTO lessons (subject, chapter_title, level, video_url, uploaded_by)
                       VALUES (%s,%s,%s,%s,%s)
                       RETURNING *""", (subject, title, level, video_url or None, user["id"]))
        lesson = cur.fetchone()
        conn.commit()
        cur.close()
    return jsonify({"lesson": api_lesson_payload(lesson)}), 201

@app.patch("/api/lessons/<int:lesson_id>")
@api_login_required({"admin", "developer", "teacher"})
def api_update_lesson(lesson_id):
    data = request.get_json(silent=True) or {}
    user = request.api_user
    title = (data.get("title") or data.get("chapter_title") or "").strip()
    video_url = video_embed_url(data.get("url") or data.get("videoUrl") or "")
    if not title:
        return api_error("Title is required.", 400, "missing_title")
    owner_clause = "AND uploaded_by=%s" if user["role"] == "teacher" else ""
    params = [title, video_url or None, lesson_id]
    if user["role"] == "teacher":
        params.append(user["id"])
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute(f"""UPDATE lessons SET chapter_title=%s, video_url=%s
                        WHERE id=%s {owner_clause}
                        RETURNING *""", params)
        lesson = cur.fetchone()
        conn.commit()
        cur.close()
    if not lesson:
        return api_error("Lesson not found.", 404, "not_found")
    return jsonify({"lesson": api_lesson_payload(lesson)})

@app.delete("/api/lessons/<int:lesson_id>")
@api_login_required({"admin", "developer", "teacher"})
def api_delete_lesson(lesson_id):
    user = request.api_user
    owner_clause = "AND uploaded_by=%s" if user["role"] == "teacher" else ""
    params = [lesson_id]
    if user["role"] == "teacher":
        params.append(user["id"])
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute(f"DELETE FROM lessons WHERE id=%s {owner_clause} RETURNING id", params)
        deleted = cur.fetchone()
        conn.commit()
        cur.close()
    if not deleted:
        return api_error("Lesson not found.", 404, "not_found")
    return jsonify({"deleted": True, "id": str(lesson_id)})

@app.get("/api/guest-videos")
def api_guest_videos():
    return jsonify({"items": []})

@app.get("/api/archive-files")
def api_archive_files():
    items = []
    for course in fetch_free_pdfs(active_only=True):
        for subject in course["subjects"]:
            for pdf in subject["pdfs"]:
                items.append({
                    "id": str(pdf["id"]),
                    "title": pdf["title"],
                    "url": pdf["preview_url"],
                    "description": subject["name"],
                    "courseId": course["level"],
                    "createdAt": pdf["created_at"].isoformat() if pdf.get("created_at") else None,
                })
    return jsonify({"items": items})

@app.get("/api/offers")
def api_offers():
    slides = fetch_offer_slides(active_only=True)
    settings = fetch_app_settings()
    return jsonify({
        "offers": [offer_slide_payload(slide) for slide in slides],
        "textSection": {
            "title": settings.get("offerTextTitle", ""),
            "body": settings.get("offerTextBody", ""),
            "active": settings.get("offerTextActive", "true") != "false",
        },
    })

@app.get("/api/visitors/online")
def api_online_visitors():
    global _visitor_settings_cache
    visitor_id = (request.headers.get("X-Visitor-Id") or "").strip()
    fallback_identity = "|".join([
        request.headers.get("X-Forwarded-For", request.remote_addr or ""),
        request.headers.get("User-Agent", ""),
    ])
    visitor_key_source = visitor_id or fallback_identity or str(random.random())
    visitor_key = hashlib.sha256(visitor_key_source.encode("utf-8")).hexdigest()
    now = time.time()
    cutoff = now - 120
    _active_visitors[visitor_key] = now
    if random.random() < 0.10:
        stale_keys = [key for key, seen_at in _active_visitors.items() if seen_at < cutoff]
        for key in stale_keys:
            _active_visitors.pop(key, None)
    online_count = sum(1 for seen_at in _active_visitors.values() if seen_at >= cutoff)

    if now - _visitor_settings_cache.get("loaded_at", 0) > 15:
        try:
            with db() as conn:
                cur = conn.cursor()
                cur.execute("""
                    SELECT key, value
                    FROM app_settings
                    WHERE key IN (
                        'visitorCountOffset',
                        'visitorAutoAddAmount',
                        'visitorAutoAddEverySeconds',
                        'visitorAutoSubtractAmount',
                        'visitorAutoSubtractEverySeconds',
                        'visitorAutomationStartedAt'
                    )
                """)
                setting_rows = cur.fetchall()
                cur.close()
            _visitor_settings_cache = {
                "loaded_at": now,
                "settings": {key: value for key, value in setting_rows},
            }
        except psycopg2.OperationalError as e:
            print("Visitor settings unavailable:", e)

    visitor_settings = _visitor_settings_cache.get("settings", {})

    def setting_int(key, default=0, minimum=0):
        try:
            return max(minimum, int(visitor_settings.get(key, str(default)) or default))
        except ValueError:
            return minimum

    offset = setting_int("visitorCountOffset")
    add_amount = setting_int("visitorAutoAddAmount")
    add_every = setting_int("visitorAutoAddEverySeconds", 60, minimum=1)
    subtract_amount = setting_int("visitorAutoSubtractAmount")
    subtract_every = setting_int("visitorAutoSubtractEverySeconds", 60, minimum=1)
    started_at = setting_int("visitorAutomationStartedAt")
    elapsed = max(0, int(now) - started_at) if started_at else 0
    auto_added = (elapsed // add_every) * add_amount if add_amount else 0
    auto_subtracted = (elapsed // subtract_every) * subtract_amount if subtract_amount else 0
    display_count = max(0, online_count + offset + auto_added - auto_subtracted)
    return jsonify({
        "online": display_count,
        "realOnline": online_count,
        "offset": offset,
        "autoAdded": auto_added,
        "autoSubtracted": auto_subtracted,
    })

@app.get("/api/results/<exam_type>/centers")
def api_result_centers(exam_type):
    if exam_type != "concours":
        return jsonify({"centers": []})
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT center_name
            FROM national_exam_results
            WHERE exam_type=%s
              AND center_name IS NOT NULL
              AND btrim(center_name) <> ''
            ORDER BY center_name ASC
        """, (exam_type,))
        centers = [row[0] for row in cur.fetchall()]
        cur.close()
    return jsonify({"centers": centers})

@app.get("/api/results/<exam_type>")
def api_search_results(exam_type):
    if exam_type not in RESULT_EXAM_TYPES:
        return api_error("Unknown exam type.", 404, "unknown_exam_type")
    query = (request.args.get("q") or "").strip()
    center = (request.args.get("center") or "").strip()
    if exam_type == "concours" and not center:
        return api_error("Centre examen is required for concours search.", 400, "missing_center")
    if not query.isdigit() and len(query) < 2:
        return jsonify({"results": []})

    with db() as conn:
        cur = dict_cursor(conn)
        if query.isdigit():
            normalized_number = normalize_candidate_number(query)
            if exam_type == "concours":
                cur.execute(f"""
                    SELECT *
                    FROM national_exam_results
                    WHERE exam_type=%s
                      AND center_name=%s
                      AND {concours_number_match_sql()}
                    ORDER BY id ASC
                    LIMIT 1
                """, (exam_type, center, normalized_number, normalized_number))
            else:
                cur.execute("""
                    SELECT *
                    FROM national_exam_results
                    WHERE exam_type=%s
                      AND COALESCE(NULLIF(regexp_replace(candidate_number, '\\D', '', 'g'), ''), '0')::bigint = %s::bigint
                    ORDER BY id ASC
                    LIMIT 1
                """, (exam_type, normalized_number))
        else:
            if exam_type == "concours":
                search_concours_by_name(cur, query, center_name=center)
            else:
                cur.execute("""
                    SELECT *
                    FROM national_exam_results
                    WHERE exam_type=%s AND full_name ILIKE %s
                    ORDER BY full_name ASC
                    LIMIT 1
                """, (exam_type, f"%{query}%"))
        rows = cur.fetchall()
        cur.close()
    return jsonify({"results": [api_result_payload(row) for row in rows]})

@app.post("/api/results/<exam_type>/upload")
@api_login_required(ADMIN_ROLES)
def api_upload_results(exam_type):
    if exam_type not in RESULT_EXAM_TYPES:
        return api_error("Unknown exam type.", 404, "unknown_exam_type")
    file = request.files.get("file")
    if not file or not file.filename:
        return api_error("Excel file is required.", 400, "missing_file")
    if not allowed(file.filename, EXCEL_EXT):
        return api_error("Only .xlsx or .xlsm files are supported.", 400, "invalid_file")

    try:
        parsed_rows = parse_results_workbook(file)
    except Exception as exc:
        return api_error(f"تعذر قراءة ملف Excel: {exc}", 400, "parse_failed")
    if not parsed_rows:
        return api_error("لم يتم العثور على نتائج قابلة للاستيراد.", 400, "empty_results")

    filename = secure_filename(file.filename) or "results.xlsx"
    try:
        upload = insert_national_results(exam_type, filename, parsed_rows, request.api_user["id"])
    except Exception as exc:
        app.logger.exception("National results API import failed")
        return api_error(f"تعذر حفظ النتائج في قاعدة البيانات: {exc}", 500, "import_failed")
    return jsonify({
        "uploaded": True,
        "rowsImported": len(parsed_rows),
        "uploadId": str(upload["id"]),
        "uploadedAt": upload["uploaded_at"].isoformat() if upload.get("uploaded_at") else None,
    }), 201

@app.get("/api/notifications")
@api_login_required()
def api_notifications():
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""SELECT id, title, body, created_at
                       FROM notifications
                       ORDER BY created_at DESC
                       LIMIT 100""")
        rows = cur.fetchall()
        cur.close()
    return jsonify({
        "notifications": [
            {
                "id": str(row["id"]),
                "title": row["title"],
                "body": row["body"],
                "createdAt": row["created_at"].isoformat() if row.get("created_at") else None,
            }
            for row in rows
        ]
    })

@app.post("/api/notifications")
@api_login_required(ADMIN_ROLES)
def api_add_notification():
    data = request.get_json(silent=True) or {}
    title = (data.get("title") or "").strip()
    body = (data.get("body") or "").strip()
    if not title or not body:
        return api_error("Title and body are required.", 400, "missing_fields")
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""INSERT INTO notifications (title, body)
                       VALUES (%s,%s)
                       RETURNING id, title, body, created_at""", (title, body))
        row = cur.fetchone()
        conn.commit()
        cur.close()
    return jsonify({
        "notification": {
            "id": str(row["id"]),
            "title": row["title"],
            "body": row["body"],
            "createdAt": row["created_at"].isoformat() if row.get("created_at") else None,
        }
    }), 201

@app.patch("/api/notifications/<int:notification_id>")
@api_login_required(ADMIN_ROLES)
def api_update_notification(notification_id):
    data = request.get_json(silent=True) or {}
    title = (data.get("title") or "").strip()
    body = (data.get("body") or "").strip()
    if not title or not body:
        return api_error("Title and body are required.", 400, "missing_fields")
    with db() as conn:
        cur = dict_cursor(conn)
        cur.execute("""UPDATE notifications
                       SET title=%s, body=%s
                       WHERE id=%s
                       RETURNING id, title, body, created_at""",
                    (title, body, notification_id))
        row = cur.fetchone()
        conn.commit()
        cur.close()
    if not row:
        return api_error("Notification not found.", 404, "not_found")
    return jsonify({
        "notification": {
            "id": str(row["id"]),
            "title": row["title"],
            "body": row["body"],
            "createdAt": row["created_at"].isoformat() if row.get("created_at") else None,
        }
    })

@app.delete("/api/notifications/<int:notification_id>")
@api_login_required(ADMIN_ROLES)
def api_delete_notification(notification_id):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM notifications WHERE id=%s", (notification_id,))
        conn.commit()
        cur.close()
    return jsonify({"deleted": True, "id": str(notification_id)})

if __name__ == "__main__":
    try:
        # Pick the outbound LAN address instead of a virtual adapter like 192.168.56.1.
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
    except Exception:
        try:
            local_ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            local_ip = "127.0.0.1"
    print("Uploads:", UP)
    print(f"Open from this PC: http://127.0.0.1:5000")
    print(f"Open from phone on same Wi-Fi: http://{local_ip}:5000")
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
