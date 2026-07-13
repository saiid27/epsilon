#!/usr/bin/env python3
import argparse
import csv
import json
import os
import re
import unicodedata
from datetime import datetime
from io import StringIO
from pathlib import Path

import psycopg2
from openpyxl import load_workbook


DEFAULT_EXCEL = Path("/Users/dehan/Downloads/FICHIER_RESULTATS_BAC_2025_53148.xlsx")
DEFAULT_DBNAME = "school_app"
EXACT_TABLE = "bac_2025_results_53148"
EXAM_TYPE = "bac-first"


def connect(dbname=None):
    database_url = os.getenv("DATABASE_URL")
    if database_url:
        if "render.com" in database_url.lower() and "sslmode=" not in database_url.lower():
            database_url += "&sslmode=require" if "?" in database_url else "?sslmode=require"
        return psycopg2.connect(database_url)
    return psycopg2.connect(
        host=os.getenv("PGHOST", "localhost"),
        port=os.getenv("PGPORT", "5432"),
        user=os.getenv("PGUSER", "postgres"),
        password=os.getenv("PGPASSWORD", "postgres"),
        dbname=dbname or os.getenv("PGDATABASE", DEFAULT_DBNAME),
    )


def ensure_database(dbname):
    if os.getenv("DATABASE_URL"):
        return
    admin = connect("postgres")
    admin.autocommit = True
    cur = admin.cursor()
    cur.execute("SELECT 1 FROM pg_database WHERE datname=%s", (dbname,))
    if not cur.fetchone():
        cur.execute(f'CREATE DATABASE "{dbname}"')
    cur.close()
    admin.close()


def normalize_identifier(value, used):
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    replacements = {
        "nodoss": "nodoss",
        "serie": "serie",
        "typec": "typec",
        "nom_fr": "nom_fr",
        "nom_ar": "nom_ar",
        "datn": "datn",
        "lieun_fr": "lieun_fr",
        "lieunn_ar": "lieunn_ar",
        "moy bac": "moy_bac",
        "decision": "decision",
        "wilaya_fr": "wilaya_fr",
        "wilaya_ar": "wilaya_ar",
        "centre examen_fr": "centre_examen_fr",
        "etablissement_fr": "etablissement_fr",
        "centre examen_ar": "centre_examen_ar",
        "etablissement_ar": "etablissement_ar",
    }
    text = replacements.get(text, text)
    text = re.sub(r"[^a-z0-9_]+", "_", text).strip("_")
    if not text:
        text = "column"
    base = text[:55]
    name = base
    index = 2
    while name in used:
        suffix = f"_{index}"
        name = f"{base[: 63 - len(suffix)]}{suffix}"
        index += 1
    used.add(name)
    return name


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def score_text(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value).replace(".", ",")
    return cell_text(value)


def read_workbook(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    used = set()
    columns = [normalize_identifier(header, used) for header in headers]
    original_headers = [cell_text(header) for header in headers]
    data = []
    for row in rows:
        values = [cell_text(value) for value in row[: len(columns)]]
        if not any(values):
            continue
        if len(values) < len(columns):
            values.extend([""] * (len(columns) - len(values)))
        data.append(values)
    wb.close()
    return original_headers, columns, data


def ensure_app_schema(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username VARCHAR(100) NOT NULL UNIQUE,
            phone VARCHAR(30) NOT NULL UNIQUE,
            password VARCHAR(64) NOT NULL,
            role VARCHAR(20) NOT NULL DEFAULT 'student',
            level VARCHAR(40),
            subject VARCHAR(80),
            status VARCHAR(20) NOT NULL DEFAULT 'pending',
            phone_verified BOOLEAN NOT NULL DEFAULT FALSE,
            payment_image VARCHAR(255),
            payment_status VARCHAR(20) DEFAULT 'pending',
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS result_uploads (
            id SERIAL PRIMARY KEY,
            exam_type VARCHAR(40) NOT NULL,
            original_filename VARCHAR(255) NOT NULL,
            rows_imported INT NOT NULL DEFAULT 0,
            uploaded_by INT REFERENCES users(id) ON DELETE SET NULL,
            uploaded_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS national_exam_results (
            id SERIAL PRIMARY KEY,
            exam_type VARCHAR(40) NOT NULL,
            candidate_number VARCHAR(80),
            full_name VARCHAR(255) NOT NULL,
            birth_place VARCHAR(160),
            birth_date VARCHAR(80),
            wilaya VARCHAR(160),
            moughataa VARCHAR(160),
            center_name VARCHAR(255),
            score VARCHAR(80),
            decision VARCHAR(160),
            rank VARCHAR(80),
            raw_data JSONB NOT NULL DEFAULT '{}'::jsonb,
            upload_id INT REFERENCES result_uploads(id) ON DELETE SET NULL,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_national_results_exam_number
        ON national_exam_results (exam_type, candidate_number)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_national_results_exam_name
        ON national_exam_results (exam_type, lower(full_name))
    """)


def create_exact_table(cur, columns):
    column_sql = ",\n            ".join(f'"{column}" TEXT' for column in columns)
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {EXACT_TABLE} (
            id SERIAL PRIMARY KEY,
            {column_sql},
            imported_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute(f"TRUNCATE TABLE {EXACT_TABLE} RESTART IDENTITY")
    cur.execute(f'CREATE INDEX IF NOT EXISTS idx_{EXACT_TABLE}_nodoss ON {EXACT_TABLE} (nodoss)')
    cur.execute(f'CREATE INDEX IF NOT EXISTS idx_{EXACT_TABLE}_nom_fr ON {EXACT_TABLE} (lower(nom_fr))')
    cur.execute(f'CREATE INDEX IF NOT EXISTS idx_{EXACT_TABLE}_nom_ar ON {EXACT_TABLE} (nom_ar)')


def copy_exact_rows(cur, columns, rows):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerows(rows)
    buffer.seek(0)
    column_sql = ", ".join(f'"{column}"' for column in columns)
    cur.copy_expert(
        f"COPY {EXACT_TABLE} ({column_sql}) FROM STDIN WITH (FORMAT CSV)",
        buffer,
    )


def copy_app_rows(cur, original_headers, columns, rows, filename):
    cur.execute("""
        INSERT INTO result_uploads (exam_type, original_filename, rows_imported, uploaded_by)
        VALUES (%s,%s,%s,NULL)
        RETURNING id
    """, (EXAM_TYPE, filename, len(rows)))
    upload_id = cur.fetchone()[0]
    cur.execute("DELETE FROM national_exam_results WHERE exam_type=%s", (EXAM_TYPE,))

    index = {name: position for position, name in enumerate(columns)}
    buffer = StringIO()
    writer = csv.writer(buffer)
    for row in rows:
        raw_data = {
            header: row[position]
            for position, header in enumerate(original_headers)
            if position < len(row)
        }
        writer.writerow((
            EXAM_TYPE,
            row[index["nodoss"]],
            row[index["nom_fr"]],
            row[index["lieun_fr"]],
            row[index["datn"]],
            row[index["wilaya_fr"]],
            "",
            row[index["centre_examen_fr"]],
            score_text(row[index["moy_bac"]]),
            row[index["decision"]],
            "",
            json.dumps(raw_data, ensure_ascii=False),
            upload_id,
        ))
    buffer.seek(0)
    cur.copy_expert("""
        COPY national_exam_results
            (exam_type, candidate_number, full_name, birth_place, birth_date,
             wilaya, moughataa, center_name, score, decision, rank, raw_data, upload_id)
        FROM STDIN WITH (FORMAT CSV)
    """, buffer)
    return upload_id


def main():
    parser = argparse.ArgumentParser(description="Import BAC 2025 Excel results into PostgreSQL.")
    parser.add_argument("--file", default=str(DEFAULT_EXCEL))
    parser.add_argument("--dbname", default=os.getenv("PGDATABASE", DEFAULT_DBNAME))
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"Excel file not found: {path}")

    ensure_database(args.dbname)
    original_headers, columns, rows = read_workbook(path)
    if not rows:
        raise SystemExit("No data rows found in the Excel file.")

    conn = connect(args.dbname)
    try:
        cur = conn.cursor()
        ensure_app_schema(cur)
        create_exact_table(cur, columns)
        copy_exact_rows(cur, columns, rows)
        upload_id = copy_app_rows(cur, original_headers, columns, rows, path.name)
        cur.execute(f"SELECT COUNT(*) FROM {EXACT_TABLE}")
        exact_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM national_exam_results WHERE exam_type=%s", (EXAM_TYPE,))
        app_count = cur.fetchone()[0]
        conn.commit()
        cur.close()
    finally:
        conn.close()

    print(f"Imported {exact_count} rows into {EXACT_TABLE}.")
    print(f"Imported {app_count} searchable rows into national_exam_results.")
    print(f"Result upload id: {upload_id}.")
    print(f"Columns ({len(columns)}): {', '.join(columns)}")


if __name__ == "__main__":
    main()
